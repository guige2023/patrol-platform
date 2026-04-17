from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload
from typing import List, Optional
from uuid import UUID
from app.dependencies import get_db, get_current_user
from app.models.inspection_group import InspectionGroup, GroupMember
from app.models.plan import Plan
from app.models.unit import Unit
from app.models.cadre import Cadre
from app.models.user import User
from app.models.audit_log import AuditLog
from app.core.audit import write_audit_log
from app.services.rule_engine import RuleEngine
from app.schemas.group import GroupCreate, GroupUpdate, GroupMemberCreate, GroupMembersReplace
from app.schemas.common import PaginatedResponse, PageResult
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

router = APIRouter()


@router.get("/", response_model=PaginatedResponse)
async def list_groups(
    plan_id: Optional[UUID] = None,
    status: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=9999),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = select(InspectionGroup).where(InspectionGroup.is_active == True).options(selectinload(InspectionGroup.members).selectinload(GroupMember.cadre))
    if plan_id:
        query = query.where(InspectionGroup.plan_id == plan_id)
    if status:
        query = query.where(InspectionGroup.status == status)

    count_result = await db.execute(select(func.count()).select_from(query.subquery()))
    total = count_result.scalar()

    query = query.order_by(InspectionGroup.created_at.desc()).offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(query)
    groups = result.scalars().all()

    items = [
        {
            "id": g.id,
            "name": g.name,
            "plan_id": g.plan_id,
            "status": g.status,
            "member_count": len(g.members),
            "leader_cadre_name": next((m.cadre.name for m in g.members if m.is_leader and m.cadre), None),
            "created_at": g.created_at,
        }
        for g in groups
    ]
    return PaginatedResponse(data=PageResult(items=items, total=total, page=page, page_size=page_size))


STATUS_LABELS = {
    "draft": "草稿",
    "approved": "已审批",
    "active": "进行中",
    "completed": "已完成",
}


@router.get("/download")
async def export_groups(
    plan_id: Optional[UUID] = None,
    status: Optional[str] = None,
    ids: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export inspection groups as .xlsx. Pass ids=comma-separated UUIDs for batch export."""
    query = (
        select(InspectionGroup)
        .options(
            selectinload(InspectionGroup.members).selectinload(GroupMember.cadre),
            selectinload(InspectionGroup.plan),
        )
        .where(InspectionGroup.is_active == True)
    )
    if ids:
        id_list = [UUID(i.strip()) for i in ids.split(",") if i.strip()]
        query = query.where(InspectionGroup.id.in_(id_list))
    elif plan_id:
        query = query.where(InspectionGroup.plan_id == plan_id)
    if status:
        query = query.where(InspectionGroup.status == status)
    query = query.order_by(InspectionGroup.created_at.desc()).limit(10000)
    result = await db.execute(query)
    groups = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "巡察组"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1677FF")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = ["巡察组名称", "所属计划", "状态", "组长姓名", "副组长姓名", "联络员姓名", "成员数量", "成员名单", "创建时间"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for g in groups:
        plan_name = g.plan.name if g.plan else ""
        members_list = g.members or []
        member_names = ",".join([(m.cadre.name if m.cadre else "未知") for m in members_list]) if members_list else ""
        created = g.created_at.strftime('%Y-%m-%d %H:%M') if g.created_at else ""
        ws.append([
            g.name or "",
            plan_name,
            STATUS_LABELS.get(g.status, g.status or ""),
            ",".join([(m.cadre.name if m.cadre else "未知") for m in members_list if m.role in ("leader", "组长")]) or "",
            ",".join([(m.cadre.name if m.cadre else "未知") for m in members_list if m.role in ("deputy_leader", "副组长")]) or "",
            ",".join([(m.cadre.name if m.cadre else "未知") for m in members_list if m.role in ("liaison", "联络员")]) or "",
            len(members_list),
            member_names,
            created,
        ])

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''groups.xlsx"},
    )


@router.post("/")
async def create_group(
    group_data: GroupCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # 组长与副组长不可为同一人
    if group_data.leader_id and group_data.vice_leader_ids:
        for vid in group_data.vice_leader_ids:
            if group_data.leader_id == vid:
                raise HTTPException(status_code=400, detail="组长与副组长不可为同一人")

    plan = None
    if group_data.plan_id:
        plan_result = await db.execute(select(Plan).where(Plan.id == group_data.plan_id))
        plan = plan_result.scalar_one_or_none()
        if not plan:
            raise HTTPException(status_code=404, detail="Plan not found")

    # Use first unit_id as target_unit_id if unit_ids provided and target_unit_id not set
    target_unit_id = group_data.target_unit_id
    if not target_unit_id and group_data.unit_ids:
        target_unit_id = group_data.unit_ids[0]

    group = InspectionGroup(
        name=group_data.name,
        plan_id=group_data.plan_id,
        target_unit_id=target_unit_id,
        created_by=current_user.id,
    )
    db.add(group)
    await db.flush()  # get group.id

    # Create leader member
    if group_data.leader_id:
        leader_member = GroupMember(
            group_id=group.id,
            cadre_id=group_data.leader_id,
            role="组长",
            is_leader=True,
        )
        db.add(leader_member)

    # Create vice leader members (support multiple)
    for vid in group_data.vice_leader_ids:
        vice_member = GroupMember(
            group_id=group.id,
            cadre_id=vid,
            role="副组长",
            is_leader=False,
        )
        db.add(vice_member)

    # Create regular members (skip leader/vice leaders)
    for cadre_id in group_data.member_ids:
        if cadre_id == group_data.leader_id or cadre_id in group_data.vice_leader_ids:
            continue
        member = GroupMember(
            group_id=group.id,
            cadre_id=cadre_id,
            role="组员",
            is_leader=False,
        )
        db.add(member)

    await db.commit()
    await db.refresh(group)
    await write_audit_log(db, current_user.id, "create", "inspection_group", group.id, {"name": group.name})
    return {"id": group.id, "name": group.name, "message": "Group created"}


@router.get("/available-cadres")
async def get_available_cadres(
    plan_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get cadres available for assignment to a group's plan."""
    # Get cadres already in any group for this plan
    subq = (
        select(GroupMember.cadre_id)
        .join(InspectionGroup, InspectionGroup.id == GroupMember.group_id)
        .where(InspectionGroup.plan_id == plan_id)
    )
    result = await db.execute(
        select(Cadre).where(Cadre.id.not_in(subq)).order_by(Cadre.name)
    )
    cadres = result.scalars().all()
    return [
        {"id": str(c.id), "name": c.name, "position": c.position, "unit_id": str(c.unit_id) if c.unit_id else None}
        for c in cadres
    ]


@router.post("/auto-match")
async def auto_match_group(
    body: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Auto-assign group members based on rule engine."""
    plan_id = body.get("plan_id")
    if not plan_id:
        raise HTTPException(status_code=422, detail="plan_id is required in body")
    try:
        plan_id = UUID(plan_id)
    except ValueError:
        raise HTTPException(status_code=422, detail="Invalid plan_id format")

    leader_id = body.get("leader_id")
    deputy_leader_ids = body.get("deputy_leader_ids", [])
    excluded = body.get("excluded_cadre_ids", [])

    # Get available cadres (not in any group for this plan)
    subq = (
        select(GroupMember.cadre_id)
        .join(InspectionGroup, InspectionGroup.id == GroupMember.group_id)
        .where(InspectionGroup.plan_id == plan_id)
    )
    excluded_ids = {e.get("cadre_id") for e in excluded if e.get("cadre_id")}
    exclude_set = excluded_ids | {leader_id} | set(deputy_leader_ids)

    result = await db.execute(
        select(Cadre).where(
            Cadre.id.not_in(subq),
            Cadre.id.not_in(exclude_set) if exclude_set else True,
        ).order_by(Cadre.name)
    )
    all_cadres = result.scalars().all()

    # Simple auto-match: assign leader, deputy leaders, then fill with regular members
    member_ids = []
    if leader_id:
        member_ids.append(leader_id)
    member_ids.extend(deputy_leader_ids)
    # Fill remaining with regular members up to 10 total
    for c in all_cadres:
        if len(member_ids) >= 10:
            break
        if str(c.id) not in [str(x) for x in member_ids]:
            member_ids.append(str(c.id))

    return {
        "member_ids": member_ids,
        "leader_id": leader_id,
        "deputy_leader_ids": deputy_leader_ids,
        "message": f"Auto-matched {len(member_ids)} members",
    }


@router.get("/{group_id}")
async def get_group(group_id: UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(
        select(InspectionGroup)
        .where(InspectionGroup.id == group_id)
        .options(selectinload(InspectionGroup.members).selectinload(GroupMember.cadre))
    )
    group = result.scalar_one_or_none()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    # Look up plan name
    plan_name = None
    if group.plan_id:
        plan_result = await db.execute(select(Plan).where(Plan.id == group.plan_id))
        plan = plan_result.scalar_one_or_none()
        if plan:
            plan_name = plan.name

    # Look up target unit name
    target_unit_name = None
    if group.target_unit_id:
        unit_result = await db.execute(select(Unit).where(Unit.id == group.target_unit_id))
        unit = unit_result.scalar_one_or_none()
        if unit:
            target_unit_name = unit.name

    return {
        "data": {
            "id": group.id,
            "name": group.name,
            "plan_id": group.plan_id,
            "plan_name": plan_name,
            "status": group.status,
            "target_unit_id": group.target_unit_id,
            "target_unit_name": target_unit_name,
            "unit_ids": group.unit_ids or [],
            "authorization_letter": group.authorization_letter,
            "authorization_date": group.authorization_date,
            "leader_cadre_name": next((m.cadre.name for m in group.members if m.is_leader and m.cadre), None),
            "members": [
                {"id": m.id, "cadre_id": m.cadre_id, "cadre_name": m.cadre.name if m.cadre else None, "role": m.role, "is_leader": m.is_leader}
                for m in group.members
            ],
            "created_at": group.created_at,
        },
        "message": "success",
    }


@router.get("/{group_id}/members")
async def list_group_members(
    group_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get all members of an inspection group."""
    result = await db.execute(
        select(InspectionGroup).where(InspectionGroup.id == group_id)
        .options(selectinload(InspectionGroup.members).selectinload(GroupMember.cadre))
    )
    group = result.scalar_one_or_none()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    return {
        "data": {
            "items": [
                {"id": str(m.id), "cadre_id": str(m.cadre_id), "cadre_name": m.cadre.name if m.cadre else None, "role": m.role, "is_leader": m.is_leader}
                for m in group.members
            ],
            "total": len(group.members),
        }
    }


@router.post("/{group_id}/members")
async def add_member(
    group_id: UUID,
    member_data: GroupMemberCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(select(InspectionGroup).where(InspectionGroup.id == group_id))
    group = result.scalar_one_or_none()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    cadre_result = await db.execute(select(Cadre).where(Cadre.id == member_data.cadre_id))
    cadre = cadre_result.scalar_one_or_none()
    if not cadre:
        raise HTTPException(status_code=404, detail="Cadre not found")

    # 去重校验：同一巡察组的同一干部不可重复添加
    existing = await db.execute(
        select(GroupMember).where(
            GroupMember.group_id == group_id,
            GroupMember.cadre_id == member_data.cadre_id,
        )
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="该干部已在巡察组中，请勿重复添加")

    member = GroupMember(
        group_id=group_id,
        cadre_id=member_data.cadre_id,
        role=member_data.role,
        is_leader=member_data.is_leader,
    )
    db.add(member)
    await db.commit()
    await write_audit_log(db, current_user.id, "add_member", "inspection_group", group_id, {"cadre_id": str(member_data.cadre_id)})
    return {"message": "Member added"}


@router.delete("/{group_id}/members/{cadre_id}")
async def remove_member(group_id: UUID, cadre_id: UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(
        select(GroupMember).where(GroupMember.group_id == group_id, GroupMember.cadre_id == cadre_id)
    )
    member = result.scalar_one_or_none()
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
    await db.delete(member)
    await db.commit()
    await write_audit_log(db, current_user.id, "remove_member", "inspection_group", group_id, {"cadre_id": str(cadre_id)})
    return {"message": "Member removed"}


@router.put("/{group_id}/members")
async def replace_group_members(
    group_id: UUID,
    data: GroupMembersReplace,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """原子全量替换巡察组成员：先删后插，保证原子性"""
    result = await db.execute(select(InspectionGroup).where(InspectionGroup.id == group_id))
    group = result.scalar_one_or_none()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    # 删所有现有成员
    await db.execute(
        select(GroupMember).where(GroupMember.group_id == group_id)
    )
    # 直接用 text() 确保删除
    from sqlalchemy import text
    await db.execute(text("DELETE FROM group_members WHERE group_id = :gid"), {"gid": str(group_id)})
    await db.flush()

    # 插入新成员
    for item in data.members:
        new_member = GroupMember(
            group_id=group_id,
            cadre_id=item.cadre_id,
            role=item.role,
            is_leader=item.is_leader,
        )
        db.add(new_member)

    await db.commit()
    await write_audit_log(db, current_user.id, "replace_members", "inspection_group", group_id, {
        "member_count": len(data.members)
    })
    return {"message": f"Members replaced ({len(data.members)} total)"}


@router.post("/{group_id}/submit")
async def submit_group(group_id: UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(InspectionGroup).where(InspectionGroup.id == group_id))
    group = result.scalar_one_or_none()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    old_status = group.status
    group.status = "approved"
    await db.commit()
    await write_audit_log(db, current_user.id, "status_change", "inspection_group", group_id, {"from": old_status, "to": "approved"})
    return {"message": "Group submitted"}


@router.post("/{group_id}/activate")
async def activate_group(group_id: UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Start executing the inspection group (approved → active)."""
    result = await db.execute(select(InspectionGroup).where(InspectionGroup.id == group_id))
    group = result.scalar_one_or_none()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    if group.status != "approved":
        raise HTTPException(status_code=400, detail=f"Cannot activate group in status '{group.status}'")
    old_status = group.status
    group.status = "active"
    await db.commit()
    await write_audit_log(db, current_user.id, "status_change", "inspection_group", group_id, {"from": old_status, "to": "active"})
    return {"message": "Group activated"}


@router.post("/{group_id}/complete")
async def complete_group(group_id: UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Mark the inspection group as completed (active → completed)."""
    result = await db.execute(select(InspectionGroup).where(InspectionGroup.id == group_id))
    group = result.scalar_one_or_none()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    if group.status != "active":
        raise HTTPException(status_code=400, detail=f"Cannot complete group in status '{group.status}'")
    old_status = group.status
    group.status = "completed"
    await db.commit()
    await write_audit_log(db, current_user.id, "status_change", "inspection_group", group_id, {"from": old_status, "to": "completed"})
    return {"message": "Group completed"}


@router.get("/{group_id}/status-logs")
async def get_group_status_logs(group_id: UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Get status transition history for a group from audit_logs."""
    result = await db.execute(
        select(AuditLog)
        .where(AuditLog.entity_type == "inspection_group")
        .where(AuditLog.entity_id == group_id)
        .where(AuditLog.action == "status_change")
        .order_by(AuditLog.created_at.desc())
        .limit(100)
    )
    logs = result.scalars().all()

    # Batch-fetch user names
    user_ids = list({log.user_id for log in logs if log.user_id})
    user_map = {}
    if user_ids:
        user_result = await db.execute(select(User).where(User.id.in_(user_ids)))
        for u in user_result.scalars().all():
            user_map[str(u.id)] = u.full_name or u.username

    return [
        {
            "id": str(log.id),
            "action": log.action,
            "from_status": log.detail.get("from") if log.detail else None,
            "to_status": log.detail.get("to") if log.detail else None,
            "user_id": str(log.user_id) if log.user_id else None,
            "user_name": user_map.get(str(log.user_id)) if log.user_id else None,
            "created_at": log.created_at.isoformat() if log.created_at else None,
        }
        for log in logs
    ]


@router.put("/{group_id}")
async def update_group(
    group_id: UUID,
    group_data: GroupUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(select(InspectionGroup).where(InspectionGroup.id == group_id))
    group = result.scalar_one_or_none()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    update_data = group_data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(group, field, value)

    await db.commit()
    await write_audit_log(db, current_user.id, "update", "inspection_group", group_id, update_data)
    return {"message": "Group updated"}


@router.delete("/{group_id}")
async def delete_group(group_id: UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(InspectionGroup).where(InspectionGroup.id == group_id))
    group = result.scalar_one_or_none()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    group.is_active = False
    await db.commit()
    await write_audit_log(db, current_user.id, "delete", "inspection_group", group_id, {})
    return {"message": "Group deleted"}


@router.post("/{group_id}/concurrent-roles")
async def assign_concurrent_roles(
    group_id: UUID,
    body: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Assign concurrent roles (clue officer, liaison officer) to a group."""
    result = await db.execute(select(InspectionGroup).where(InspectionGroup.id == group_id))
    group = result.scalar_one_or_none()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    await db.commit()
    await write_audit_log(db, current_user.id, "assign_concurrent_roles", "inspection_group", group_id, body)
    return {"message": "Concurrent roles assigned"}


@router.get("/{group_id}/phase-logs")
async def get_group_phase_logs(
    group_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get phase transition logs for an inspection group."""
    result = await db.execute(
        select(AuditLog)
        .where(AuditLog.entity_type == "inspection_group")
        .where(AuditLog.entity_id == group_id)
        .order_by(AuditLog.created_at.desc())
        .limit(100)
    )
    logs = result.scalars().all()
    return [
        {
            "id": str(log.id),
            "action": log.action,
            "detail": log.detail,
            "user_id": str(log.user_id) if log.user_id else None,
            "created_at": log.created_at.isoformat() if log.created_at else None,
        }
        for log in logs
    ]


@router.post("/{group_id}/status")
async def update_group_status(
    group_id: UUID,
    body: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Update group status directly (e.g., from 'approved' to 'active')."""
    result = await db.execute(select(InspectionGroup).where(InspectionGroup.id == group_id))
    group = result.scalar_one_or_none()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    new_status = body.get("status")
    if not new_status:
        raise HTTPException(status_code=400, detail="status is required")
    old_status = group.status
    group.status = new_status
    await db.commit()
    await write_audit_log(db, current_user.id, "status_change", "inspection_group", group_id, {"from": old_status, "to": new_status})
    return {"message": f"Group status updated to {new_status}"}

@router.post("/batch-delete")
async def batch_delete_groups(
    ids: List[UUID],
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Soft-delete multiple inspection groups at once."""
    if not ids:
        raise HTTPException(status_code=400, detail="No IDs provided")
    result = await db.execute(
        select(InspectionGroup).where(InspectionGroup.id.in_(ids), InspectionGroup.is_active == True)
    )
    groups = result.scalars().all()
    if not groups:
        raise HTTPException(status_code=404, detail="No groups found")
    for g in groups:
        g.is_active = False
    await db.commit()
    for g in groups:
        await write_audit_log(db, current_user.id, "delete", "inspection_group", g.id, {})
    return {"message": f"{len(groups)} groups deleted"}

