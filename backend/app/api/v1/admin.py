from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, or_, cast, String, insert, delete
from uuid import UUID
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.dependencies import get_uow, get_current_user, require_permission
from app.database import UnitOfWork
from app.models.user import User, Role, Permission, user_roles
from app.models.unit import Unit
from app.models.module_config import ModuleConfig
from app.models.rule_config import RuleConfig
from app.models.audit_log import AuditLog
from app.core.security import get_password_hash
from app.schemas.user import UserCreate, UserUpdate
from app.core.audit import write_audit_log

router = APIRouter()


@router.get("/users")
async def list_users(
    uow: UnitOfWork = Depends(get_uow),
    current_user: User = Depends(require_permission("user:read")),
):
    result = await uow.execute(select(User).order_by(User.created_at.desc()))
    users = result.scalars().all()
    return [
        {"id": u.id, "username": u.username, "email": u.email, "full_name": u.full_name, "role": u.role, "is_active": u.is_active}
        for u in users
    ]


@router.post("/users")
async def create_user(
    user_data: UserCreate,
    uow: UnitOfWork = Depends(get_uow),
    current_user: User = Depends(require_permission("user:write")),
):
    existing = await uow.execute(select(User).where(User.username == user_data.username))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Username already exists")
    user = User(
        username=user_data.username,
        email=user_data.email,
        hashed_password=get_password_hash(user_data.password),
        full_name=user_data.full_name,
        role=user_data.role,
        unit_id=user_data.unit_id,
        is_active=user_data.is_active,
    )
    uow.add(user)
    await uow.flush()

    # 同步 user_roles 关联表（直接操作关联表，避免 ORM relationship 的 greenlet 问题）
    if user_data.role:
        role_result = await uow.execute(
            select(Role).where((Role.name == user_data.role) | (Role.code == user_data.role))
        )
        role = role_result.scalar_one_or_none()
        if role:
            await uow.execute(
                insert(user_roles).values(user_id=user.id, role_id=role.id)
            )

    await uow.commit()
    await uow.refresh(user)
    await write_audit_log(uow.session, current_user.id, "create", "user", user.id, {"username": user.username})
    return {"id": user.id, "username": user.username}


@router.put("/users/{user_id}")
async def update_user(
    user_id: UUID,
    user_data: UserUpdate,
    uow: UnitOfWork = Depends(get_uow),
    current_user: User = Depends(require_permission("user:write")),
):
    result = await uow.execute(select(User).where(User.id == user_id))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    update_data = user_data.model_dump(exclude_unset=True)
    if "password" in update_data and update_data["password"]:
        user.hashed_password = get_password_hash(update_data.pop("password"))
    for key, value in update_data.items():
        setattr(user, key, value)

    # 同步 user_roles 关联表
    if "role" in update_data:
        # 先删除旧关联
        await uow.execute(delete(user_roles).where(user_roles.c.user_id == user.id))
        if update_data["role"]:
            role_result = await uow.execute(
                select(Role).where((Role.name == update_data["role"]) | (Role.code == update_data["role"]))
            )
            role = role_result.scalar_one_or_none()
            if role:
                await uow.execute(
                    insert(user_roles).values(user_id=user.id, role_id=role.id)
                )

    await uow.commit()
    await write_audit_log(uow.session, current_user.id, "update", "user", user.id, {"username": user.username})
    return {"id": user.id, "username": user.username}


@router.delete("/users/{user_id}")
async def delete_user(
    user_id: UUID,
    uow: UnitOfWork = Depends(get_uow),
    current_user: User = Depends(require_permission("user:write")),
):
    result = await uow.execute(select(User).where(User.id == user_id))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    user.is_active = False
    await uow.commit()
    await write_audit_log(uow.session, current_user.id, "delete", "user", user.id, {"username": user.username})
    return {"message": "User deleted"}


# 中英文映射
ACTION_LABELS = {
    "create": "创建", "update": "更新", "delete": "删除",
    "submit": "提交", "approve": "审批", "publish": "发布",
    "transfer": "移交", "sign": "签收", "verify": "核实",
    "confirm": "确认", "draft_submit": "底稿提交", "draft_approve": "底稿审批",
    "draft_reject": "底稿退回", "draft_publish": "底稿发布",
    "update_progress": "更新进度", "reimport": "重新导入",
    "generate": "生成", "upload": "上传", "download": "下载",
}

ENTITY_TYPE_LABELS = {
    "user": "用户", "role": "角色", "unit": "单位", "cadre": "干部",
    "plan": "巡察计划", "group": "巡察组", "draft": "底稿",
    "clue": "线索", "rectification": "整改", "knowledge": "知识库",
    "document": "文档", "progress": "进度", "notification": "通知",
    "alert": "告警", "attachment": "附件",
}


@router.get("/audit-logs")
async def list_audit_logs(
    page: int = 1,
    page_size: int = 50,
    entity_type: str = None,
    search: str = None,
    uow: UnitOfWork = Depends(get_uow),
    current_user: User = Depends(require_permission("audit:read")),
):
    # JOIN with users to get full_name
    query = (
        select(AuditLog, User.full_name)
        .outerjoin(User, AuditLog.user_id == User.id)
        .order_by(AuditLog.created_at.desc())
    )
    if entity_type:
        query = query.where(AuditLog.entity_type == entity_type)
    if search:
        search_pattern = f"%{search}%"
        query = query.where(
            or_(
                User.full_name.ilike(search_pattern),
                AuditLog.entity_id.cast(String).ilike(search_pattern)
            )
        )
    count_result = await uow.execute(select(func.count()).select_from(AuditLog))
    total = count_result.scalar()
    query = query.offset((page - 1) * page_size).limit(page_size)
    result = await uow.execute(query)
    rows = result.all()
    return {
        "items": [
            {
                "id": l[0].id,
                "user_id": str(l[0].user_id) if l[0].user_id else None,
                "username": l[1] or "未知用户",
                "action": l[0].action,
                "action_label": ACTION_LABELS.get(l[0].action, l[0].action),
                "entity_type": l[0].entity_type,
                "entity_type_label": ENTITY_TYPE_LABELS.get(l[0].entity_type, l[0].entity_type),
                "entity_id": str(l[0].entity_id),
                "changes": l[0].detail or {},
                "created_at": l[0].created_at,
            }
            for l in rows
        ],
        "total": total,
    }


@router.get("/audit-logs/download")
async def export_audit_logs(
    entity_type: str = None,
    uow: UnitOfWork = Depends(get_uow),
    current_user: User = Depends(require_permission("audit:read")),
):
    """Export audit logs as .xlsx."""
    query = select(AuditLog)
    if entity_type:
        query = query.where(AuditLog.entity_type == entity_type)
    query = query.order_by(AuditLog.created_at.desc()).limit(10000)
    result = await uow.execute(query)
    logs = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "审计日志"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1677FF")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = ["操作", "实体类型", "实体ID", "详情", "IP", "时间"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for l in logs:
        detail_str = str(l.detail) if l.detail else ""
        created = l.created_at.strftime('%Y-%m-%d %H:%M:%S') if l.created_at else ""
        ws.append([
            l.action or "",
            l.entity_type or "",
            str(l.entity_id) if l.entity_id else "",
            detail_str[:200],  # truncate long details
            l.ip_address or "",
            created,
        ])

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''audit_logs.xlsx"},
    )


@router.get("/modules")
async def list_modules(
    uow: UnitOfWork = Depends(get_uow),
    current_user: User = Depends(require_permission("module:read")),
):
    result = await uow.execute(select(ModuleConfig))
    modules = result.scalars().all()
    return modules


@router.put("/modules/{module_id}")
async def update_module(
    module_id: UUID,
    data: dict = None,
    uow: UnitOfWork = Depends(get_uow),
    current_user: User = Depends(require_permission("module:write")),
):
    """Update module config. Accepts body with is_enabled and optional config."""
    result = await uow.execute(select(ModuleConfig).where(ModuleConfig.id == module_id))
    module = result.scalar_one_or_none()
    if not module:
        raise HTTPException(status_code=404, detail="Module not found")
    if data is None:
        raise HTTPException(status_code=400, detail="Request body required")
    is_enabled = data.get("is_enabled")
    config = data.get("config")
    if is_enabled is not None:
        module.is_enabled = is_enabled
    if config is not None:
        module.config = config
    await uow.commit()
    return {"message": "Module updated"}


# ─── Roles ────────────────────────────────────────────────────────────────────

@router.get("/roles")
async def list_roles(
    uow: UnitOfWork = Depends(get_uow),
    current_user: User = Depends(require_permission("role:read")),
):
    """List all roles."""
    result = await uow.execute(select(Role).order_by(Role.created_at.desc()))
    roles = result.scalars().all()
    return [
        {
            "id": r.id,
            "name": r.name,
            "code": r.code,
            "description": r.description,
            "is_active": r.is_active,
            "permissions": r.permissions or [],
            "created_at": r.created_at,
        }
        for r in roles
    ]


@router.post("/roles")
async def create_role(
    data: dict,
    uow: UnitOfWork = Depends(get_uow),
    current_user: User = Depends(require_permission("role:write")),
):
    """Create a new role."""
    name = data.get("name")
    code = data.get("code")
    if not name or not code:
        raise HTTPException(status_code=400, detail="name and code are required")
    existing = await uow.execute(select(Role).where(Role.code == code))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Role code already exists")
    role = Role(
        name=name,
        code=code,
        description=data.get("description", ""),
        permissions=data.get("permissions", []),
        is_active=data.get("is_active", True),
    )
    uow.add(role)
    await uow.commit()
    await uow.refresh(role)
    await write_audit_log(uow.session, current_user.id, "create", "role", role.id, {"name": role.name})
    return {"id": role.id, "name": role.name}


@router.put("/roles/{role_id}")
async def update_role(
    role_id: UUID,
    data: dict,
    uow: UnitOfWork = Depends(get_uow),
    current_user: User = Depends(require_permission("role:write")),
):
    """Update a role."""
    result = await uow.execute(select(Role).where(Role.id == role_id))
    role = result.scalar_one_or_none()
    if not role:
        raise HTTPException(status_code=404, detail="Role not found")
    if "name" in data:
        role.name = data["name"]
    if "description" in data:
        role.description = data["description"]
    if "permissions" in data:
        role.permissions = data["permissions"]
    if "is_active" in data:
        role.is_active = data["is_active"]
    await uow.commit()
    await write_audit_log(uow.session, current_user.id, "update", "role", role.id, {"name": role.name})
    return {"id": role.id, "name": role.name}


@router.delete("/roles/{role_id}")
async def delete_role(
    role_id: UUID,
    uow: UnitOfWork = Depends(get_uow),
    current_user: User = Depends(require_permission("role:write")),
):
    """Soft-delete a role."""
    result = await uow.execute(select(Role).where(Role.id == role_id))
    role = result.scalar_one_or_none()
    if not role:
        raise HTTPException(status_code=404, detail="Role not found")
    role.is_active = False
    await uow.commit()
    await write_audit_log(uow.session, current_user.id, "delete", "role", role.id, {"name": role.name})
    return {"message": "Role deleted"}


@router.get("/permissions")
async def list_permissions(
    uow: UnitOfWork = Depends(get_uow),
    current_user: User = Depends(require_permission("role:read")),
):
    """List all available permissions."""
    result = await uow.execute(select(Permission).order_by(Permission.code))
    perms = result.scalars().all()
    return [
        {"id": p.id, "code": p.code, "name": p.name, "description": p.description}
        for p in perms
    ]


# ============================================================
# Admin 路由别名 - 包装其他模块的端点，使前端 /admin/* 路由可用
# 注意：由于依赖注入的限制，这些端点直接实现而不是调用原始函数
# ============================================================

from app.models.field_option import FieldOption
from app.schemas.field_option import FieldOptionResponse
from app.models.system_config import SystemConfig
from pydantic import BaseModel
from typing import Optional
import os
from pathlib import Path
import json
from datetime import datetime

# 备份目录
BACKUPS_DIR = Path(__file__).parent.parent.parent.parent / "backups"

class FieldOptionCreate(BaseModel):
    field_key: str
    option_value: str
    option_label: str
    sort_order: Optional[int] = 0
    is_active: Optional[bool] = True

class FieldOptionUpdate(BaseModel):
    option_label: Optional[str] = None
    sort_order: Optional[int] = None
    is_active: Optional[bool] = None


@router.get("/field-options")
async def list_field_options_proxy(
    uow: UnitOfWork = Depends(get_uow),
    current_user: User = Depends(require_permission("field_option:read")),
):
    """获取所有字段选项配置"""
    result = await uow.execute(select(FieldOption).order_by(FieldOption.sort_order, FieldOption.field_key))
    options = result.scalars().all()
    return [FieldOptionResponse.from_model(o) for o in options]


@router.put("/field-options/{field_key}")
async def update_field_options_proxy(
    field_key: str,
    data: dict,
    uow: UnitOfWork = Depends(get_uow),
    current_user: User = Depends(require_permission("field_option:write")),
):
    """更新字段选项"""
    result = await uow.execute(select(FieldOption).where(FieldOption.field_key == field_key))
    option = result.scalars().first()
    if not option:
        raise HTTPException(status_code=404, detail="Field option not found")
    
    if 'option_label' in data:
        option.option_label = data['option_label']
    if 'sort_order' in data:
        option.sort_order = data['sort_order']
    if 'is_active' in data:
        option.is_active = data['is_active']
    
    await uow.commit()
    await uow.refresh(option)
    return FieldOptionResponse.from_model(option)


@router.get("/system-configs")
async def get_system_configs_proxy(
    uow: UnitOfWork = Depends(get_uow),
    current_user: User = Depends(require_permission("system_config:read")),
):
    """获取所有系统配置"""
    result = await uow.execute(select(SystemConfig))
    configs = result.scalars().all()
    return {c.key: c.value for c in configs}


@router.put("/system-configs/{config_key}")
async def update_system_config_proxy(
    config_key: str,
    data: dict,
    uow: UnitOfWork = Depends(get_uow),
    current_user: User = Depends(require_permission("system_config:write")),
):
    """更新系统配置"""
    result = await uow.execute(select(SystemConfig).where(SystemConfig.key == config_key))
    config = result.scalars().first()
    if not config:
        raise HTTPException(status_code=404, detail="Config not found")

    if 'value' in data:
        config.value = data['value']

    await uow.commit()
    await uow.refresh(config)
    return {"key": config.key, "value": config.value}


@router.get("/backup")
async def list_backups_proxy(
    current_user: User = Depends(require_permission("backup:read")),
):
    """获取备份列表"""
    backups = []
    if not BACKUPS_DIR.exists():
        return {"backups": []}

    for fname in sorted(BACKUPS_DIR.iterdir()):
        if fname.suffix == ".zip":
            stat = fname.stat()
            meta_file = fname.with_suffix(".meta.json")
            if meta_file.exists():
                with open(meta_file, "r", encoding="utf-8") as f:
                    meta = json.load(f)
                meta["size"] = stat.st_size
            else:
                meta = {
                    "filename": fname.name,
                    "timestamp": datetime.fromtimestamp(stat.st_mtime).isoformat(),
                    "type": "unknown",
                    "size": stat.st_size,
                }
            backups.append(meta)

    return {"backups": sorted(backups, key=lambda x: x.get("timestamp", ""), reverse=True)}


@router.get("/backup/settings")
async def get_backup_settings_proxy(
    current_user: User = Depends(require_permission("backup:read")),
):
    """获取备份设置"""
    return {
        "auto_backup_enabled": False,
        "auto_backup_interval_hours": 24,
        "max_backups_to_keep": 10,
        "backup_types": ["manual", "auto"],
    }


@router.put("/backup/settings")
async def update_backup_settings_proxy(
    settings_update: dict,
    current_user: User = Depends(require_permission("backup:write")),
):
    """更新备份设置"""
    await write_audit_log(uow.session, current_user.id, "update", "backup_settings", None, settings_update)
    return {"message": "Settings updated"}
