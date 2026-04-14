from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from typing import List, Optional
from uuid import UUID
import io, codecs
from app.dependencies import get_db, get_current_user
from app.models.cadre import Cadre
from app.models.user import User
from app.schemas.cadre import CadreCreate, CadreUpdate, CadreResponse
from app.schemas.common import PaginatedResponse, PageResult
from app.core.audit import write_audit_log
from app.core.encryption import encrypt_field, decrypt_field, mask_id_card
import json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

router = APIRouter()


@router.get("/", response_model=PaginatedResponse[CadreResponse])
async def list_cadres(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=9999),
    name: Optional[str] = None,
    unit_id: Optional[UUID] = None,
    tags: Optional[str] = None,
    is_available: Optional[bool] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = select(Cadre).where(Cadre.is_active == True)
    if name:
        query = query.where(Cadre.name.ilike(f"%{name}%"))
    if unit_id:
        query = query.where(Cadre.unit_id == unit_id)
    if is_available is not None:
        query = query.where(Cadre.is_available == is_available)
    
    count_result = await db.execute(select(func.count()).select_from(query.subquery()))
    total = count_result.scalar()
    
    query = query.order_by(Cadre.created_at.desc()).offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(query)
    items = result.scalars().all()
    
    return PaginatedResponse(
        data=PageResult(items=items, total=total, page=page, page_size=page_size)
    )


@router.get("/export")
async def export_cadres(
    name: Optional[str] = None,
    unit_id: Optional[UUID] = None,
    is_available: Optional[bool] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export cadres as .xlsx (max 10000 rows)."""
    query = select(Cadre).where(Cadre.is_active == True)
    if name:
        query = query.where(Cadre.name.ilike(f"%{name}%"))
    if unit_id:
        query = query.where(Cadre.unit_id == unit_id)
    if is_available is not None:
        query = query.where(Cadre.is_available == is_available)
    query = query.order_by(Cadre.created_at.desc()).limit(10000)
    result = await db.execute(query)
    cadres = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "干部人才"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1677FF")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = ["姓名", "性别", "出生日期", "民族", "籍贯", "政治面貌", "学历", "学位",
               "职务", "职级", "类别", "所属单位", "标签", "简历", "是否可用", "创建时间"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for c in cadres:
        created = c.created_at.strftime("%Y-%m-%d %H:%M") if c.created_at else ""
        if isinstance(c.tags, dict):
            tags_str = c.tags.get("熟悉领域", "") or ""
        elif isinstance(c.tags, list):
            tags_str = ",".join(c.tags)
        else:
            tags_str = str(c.tags) if c.tags else ""
        ws.append([
            c.name or "", c.gender or "",
            str(c.birth_date) if c.birth_date else "",
            c.ethnicity or "", c.native_place or "", c.political_status or "",
            c.education or "", c.degree or "",
            c.position or "", c.rank or "", c.category or "",
            "",  # unit_name would need a join
            tags_str, c.resume or "",
            "是" if c.is_available else "否",
            created,
        ])

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''cadres_export.xlsx"},
    )


@router.get("/template")
async def download_cadre_template(
    current_user: User = Depends(get_current_user),
):
    """Download cadre import template (.xlsx)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "干部人才导入模板"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1677FF")
    header_align = Alignment(horizontal="center", vertical="center")
    note_fill = PatternFill("solid", fgColor="FFF7E6")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = ["姓名*", "性别", "出生日期", "民族", "籍贯", "政治面貌", "学历", "学位",
               "职务(选填)", "职级(选填)", "类别(选填)", "标签(逗号分隔)", "简历", "工作经历", "是否可用"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    notes = [
        "必填", "男/女", "如 1990-01-01", "如 汉", "如 浙江杭州", "中共党员/群众等",
        "博士研究生/硕士研究生/大学本科/大学专科/中专/高中/初中及以下",
        "博士学位/硕士学位/学士学位", "职务（见字段配置）", "职级文字",
        "纪检监察干部/审计干部/财务干部/综合干部/后备干部",
        "多个用逗号分隔", "个人简历", "工作经历", "是/否，默认为是",
    ]
    ws.append(notes)
    for cell in ws[2]:
        cell.fill = note_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        cell.font = Font(size=10, color="999999")

    sample = [
        ["张三", "男", "1985-03-15", "汉", "浙江杭州", "中共党员", "大学本科", "学士学位",
         "正处级", "一级调研员", "纪检监察干部", "优秀干部,业务骨干", "2008年毕业于XX大学",
         "2008-2012 XX局科员；2012-2018 XX处副处长", "是"],
        ["李四", "女", "1990-07-20", "汉", "江苏南京", "中共党员", "硕士研究生", "硕士学位",
         "", "", "后备干部", "新提拔", "2015年毕业于XX大学",
         "2015-2020 XX中心科员；2020至今 XX处一级主任科员", "是"],
    ]
    for row in sample:
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[2].height = 36
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 6, 45)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''cadre_template.xlsx"},
    )



@router.get("/{cadre_id}", response_model=CadreResponse)
async def get_cadre(cadre_id: UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(Cadre).where(Cadre.id == cadre_id))
    cadre = result.scalar_one_or_none()
    if not cadre:
        raise HTTPException(status_code=404, detail="Cadre not found")
    return cadre


@router.post("/", response_model=CadreResponse, status_code=201)
async def create_cadre(cadre_data: CadreCreate, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    data = cadre_data.model_dump()
    if data.get("id_card_encrypted"):
        data["id_card_encrypted"] = encrypt_field(data["id_card_encrypted"])
    cadre = Cadre(**data)
    db.add(cadre)
    await db.commit()
    await db.refresh(cadre)
    await write_audit_log(db, current_user.id, "create", "cadre", cadre.id, {"name": cadre.name})
    return cadre


@router.put("/{cadre_id}", response_model=CadreResponse)
async def update_cadre(cadre_id: UUID, cadre_data: CadreUpdate, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(Cadre).where(Cadre.id == cadre_id))
    cadre = result.scalar_one_or_none()
    if not cadre:
        raise HTTPException(status_code=404, detail="Cadre not found")
    data = cadre_data.model_dump(exclude_unset=True)
    if data.get("id_card_encrypted"):
        data["id_card_encrypted"] = encrypt_field(data["id_card_encrypted"])
    for key, value in data.items():
        setattr(cadre, key, value)
    await db.commit()
    await db.refresh(cadre)
    await write_audit_log(db, current_user.id, "update", "cadre", cadre.id, {"name": cadre.name})
    return cadre


@router.delete("/{cadre_id}")
async def delete_cadre(cadre_id: UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(Cadre).where(Cadre.id == cadre_id))
    cadre = result.scalar_one_or_none()
    if not cadre:
        raise HTTPException(status_code=404, detail="Cadre not found")
    cadre.is_active = False
    await db.commit()
    await write_audit_log(db, current_user.id, "delete", "cadre", cadre.id, {"name": cadre.name})
    return {"message": "Cadre deleted"}


@router.get("/{cadre_id}/id-card/masked")
async def get_masked_id_card(cadre_id: UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(Cadre).where(Cadre.id == cadre_id))
    cadre = result.scalar_one_or_none()
    if not cadre or not cadre.id_card_encrypted:
        raise HTTPException(status_code=404, detail="Cadre or ID card not found")
    decrypted = decrypt_field(cadre.id_card_encrypted)
    return {"masked": mask_id_card(decrypted)}


@router.post("/import")
async def import_cadres(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    导入干部数据（Excel .xlsx）
    必填列：name
    可选列：gender, birth_date, ethnicity, native_place, political_status,
            education, degree, unit_id, position, rank, category, tags, profile,
            resume, achievements(JSON), is_available
    """
    import json, openpyxl

    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 文件")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active

    headers = [str(h.value).strip() if h.value else "" for h in ws[1]]

    # 中文表头映射
    HEADER_ALIASES = {
        "姓名*": "name", "姓名": "name",
        "性别": "gender", "出生日期": "birth_date", "民族": "ethnicity",
        "籍贯": "native_place", "政治面貌": "political_status",
        "学历": "education", "学位": "degree",
        "职务(选填)": "position", "职务": "position",
        "职级(选填)": "rank", "职级": "rank", "职级级别": "rank",
        "人员类别": "category", "类别(选填)": "category", "类别": "category",
        "干部类别": "category",
        "标签(逗号分隔)": "tags", "标签": "tags",
        "熟悉领域": "tags",
        "简历": "resume", "工作经历": "work_history",
        "所属单位": "unit_id",
        "是否可用": "is_available",
        "简介": "profile", "个人简历": "resume",
    }

    # 标准化列名（去除*后缀和空格）
    # 标准化列名（去除*后缀和空格），支持部分匹配
    header_to_key = {src: dst for src, dst in HEADER_ALIASES.items()}
    col_map = {}
    for i, h in enumerate(headers):
        h_stripped = h.strip().rstrip("*").strip()
        key = header_to_key.get(h_stripped)
        if key is None:
            for alias_src, mapped_key in HEADER_ALIASES.items():
                if h_stripped.startswith(alias_src.strip().rstrip("*").strip()):
                    key = mapped_key
                    break
            else:
                key = h_stripped.lower()
        col_map[key] = i

    if "name" not in col_map:
        raise HTTPException(status_code=400, detail="缺少必填列: name")

    # 加载字段选项配置，构建校验规则
    from app.models.field_option import FieldOption
    field_opts_result = await db.execute(select(FieldOption))
    all_field_opts = field_opts_result.scalars().all()
    # field_key → set of valid labels
    field_option_map: dict[str, set[str]] = {}
    for fo in all_field_opts:
        field_option_map[fo.field_key] = {opt["label"] for opt in json.loads(fo.options) if isinstance(opt, dict) and opt.get("label")}

    # 需要校验的字段映射（col_key → field_key）
    validated_fields = {
        "category": "cadre_category",
        "position": "cadre_position",
        "rank": "cadre_rank",
    }

    created, skipped, errors = 0, 0, []
    first_id = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            name = row[col_map["name"]]
            if not name:
                errors.append(f"第{row_idx}行: name 为空")
                continue

            # 按姓名查重
            exist = await db.execute(select(Cadre).where(Cadre.name == str(name).strip()))
            if exist.scalar_one_or_none():
                skipped += 1
                continue

            data = {
                "name": str(name).strip(),
                "gender": _cell(row, col_map, "gender"),
                "birth_date": _date_cell(row, col_map, "birth_date"),
                "ethnicity": _cell(row, col_map, "ethnicity"),
                "native_place": _cell(row, col_map, "native_place"),
                "political_status": _cell(row, col_map, "political_status"),
                "education": _cell(row, col_map, "education"),
                "degree": _cell(row, col_map, "degree"),
                "unit_id": None,
                "position": _cell(row, col_map, "position"),
                "rank": _cell(row, col_map, "rank"),
                "category": _cell(row, col_map, "category"),
                "tags": {"熟悉领域": _cell(row, col_map, "tags")} if "tags" in col_map else {},
                "profile": _cell(row, col_map, "profile"),
                "resume": _cell(row, col_map, "resume"),
                "achievements": _json_cell(row, col_map, "achievements"),
                "is_available": _bool_cell(row, col_map, "is_available", default=True),
            }

            # 字段选项范围校验
            row_errors = []
            for col_key, field_key in validated_fields.items():
                # 只对Excel里有这列 且 field_key 有配置的情况下校验
                if col_key in col_map and field_key in field_option_map:
                    raw_val = data.get(col_key)
                    if raw_val:  # 只校验有值的
                        valid_labels = field_option_map[field_key]
                        if raw_val not in valid_labels:
                            row_errors.append(
                                f"[{col_key}]='{raw_val}' 不在系统允许的范围内，可选值: {', '.join(sorted(valid_labels))}"
                            )
            if row_errors:
                errors.append(f"第{row_idx}行 [{name}]: " + "；".join(row_errors))
                continue

            cadre = Cadre(**{k: v for k, v in data.items() if v is not None})
            db.add(cadre)
            await db.flush()
            if first_id is None:
                first_id = cadre.id
            created += 1
        except Exception as e:
            errors.append(f"第{row_idx}行: {str(e)}")

    await db.commit()
    # 只要有校验错误就返回详细提示（不抛异常，由前端弹窗展示）
    validation_errors = [e for e in errors if any(k in e for k in ["不在系统允许的范围内", "不在可选值范围内"])]
    if validation_errors:
        raise HTTPException(
            status_code=400,
            detail={
                "message": f"有 {len(validation_errors)} 条数据校验不通过，请修改后重新导入",
                "errors": validation_errors,
                "created": created,
                "skipped": skipped,
            }
        )

    if errors and created == 0:
        raise HTTPException(status_code=400, detail=f"导入失败: {'; '.join(errors[:5])}")

    if first_id is not None:
        await write_audit_log(db, current_user.id, "import", "cadre", first_id, {
            "created": created, "skipped": skipped
        })
    return {
        "message": f"导入完成：新增 {created} 条，{skipped} 条因姓名重复被跳过",
        "detail": f"{skipped}条记录因姓名重复被跳过，可前往列表手动编辑覆盖",
        "created": created,
        "skipped": skipped,
        "errors": errors[:10],
    } if created > 0 else {
        "message": f"无新数据导入（{skipped}条姓名重复）",
        "detail": f"{skipped}条记录因姓名重复被跳过，可前往列表手动编辑覆盖",
        "created": 0,
        "skipped": skipped,
        "errors": errors[:10],
    }


def _cell(row, col_map, key):
    if key not in col_map:
        return None
    v = row[col_map[key]]
    return str(v).strip() if v is not None else None


def _date_cell(row, col_map, key):
    from datetime import date
    v = _cell(row, col_map, key)
    if not v:
        return None
    try:
        return date.fromisoformat(str(v).strip())
    except (ValueError, TypeError):
        return None


def _bool_cell(row, col_map, key, default=False):
    v = _cell(row, col_map, key)
    if v is None:
        return default
    return str(v).lower() in ("true", "1", "是", "yes")


def _list_cell(row, col_map, key):
    v = _cell(row, col_map, key)
    if not v:
        return []
    return [s.strip() for s in str(v).split(",") if s.strip()]


def _json_cell(row, col_map, key):
    import json
    v = _cell(row, col_map, key)
    if not v:
        return None
    try:
        return json.loads(v)
    except (json.JSONDecodeError, TypeError):
        return {"raw": v}

