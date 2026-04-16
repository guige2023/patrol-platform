from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import FileResponse, Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from typing import Optional
from uuid import UUID
from datetime import datetime
from pathlib import Path
import aiofiles
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.dependencies import get_db, get_current_user
from app.models.document import Document, DocumentType
from app.models.plan import Plan
from app.models.rectification import Rectification
from app.models.user import User
from app.schemas.document import DocumentResponse, GenerateDocumentRequest, GenerateRectificationNoticeRequest
from app.schemas.common import PaginatedResponse, PageResult
from app.core.audit import write_audit_log

router = APIRouter()

DOCUMENTS_DIR = Path("/Users/guige/my_project/patrol_platform/backend/app/static/documents")
DOCUMENTS_DIR.mkdir(parents=True, exist_ok=True)


@router.get("/", response_model=PaginatedResponse[DocumentResponse])
async def list_documents(
    type: Optional[str] = None,
    plan_id: Optional[UUID] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=9999),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = select(Document).where(Document.is_active == True)
    if type:
        query = query.where(Document.type == type)
    if plan_id:
        query = query.where(Document.plan_id == plan_id)

    count_result = await db.execute(select(func.count()).select_from(query.subquery()))
    total = count_result.scalar()

    query = query.order_by(Document.generate_date.desc()).offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(query)
    items = result.scalars().all()

    return PaginatedResponse(
        data=PageResult(items=[DocumentResponse.model_validate(item) for item in items], total=total, page=page, page_size=page_size)
    )


@router.get("/{document_id}", response_model=DocumentResponse)
async def get_document(document_id: UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(Document).where(Document.id == document_id, Document.is_active == True))
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="Document not found")
    return item


@router.delete("/{document_id}")
async def delete_document(document_id: UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(Document).where(Document.id == document_id))
    doc = result.scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    doc.is_active = False
    await db.commit()
    await write_audit_log(db, current_user.id, "delete", "document", document_id, {})
    return {"message": "deleted"}


@router.get("/{document_id}/download")
async def download_document(document_id: UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(Document).where(Document.id == document_id, Document.is_active == True))
    doc = result.scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    if not doc.file_path:
        raise HTTPException(status_code=404, detail="File not found")

    file_path = Path(doc.file_path)
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found on disk")

    return FileResponse(file_path, filename=file_path.name, media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document")


@router.get("/{document_id}/preview")
async def preview_document(document_id: UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Preview document - returns file as blob."""
    result = await db.execute(select(Document).where(Document.id == document_id, Document.is_active == True))
    doc = result.scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    if not doc.file_path:
        raise HTTPException(status_code=404, detail="File not found")

    file_path = Path(doc.file_path)
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found on disk")

    async with aiofiles.open(file_path, "rb") as f:
        content = await f.read()

    return Response(content=content, media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document")


@router.post("/generate")
async def generate_document(
    request: GenerateDocumentRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Generate a document from plan (巡察公告, 成立通知, 部署会通知, 反馈意见)."""
    plan_result = await db.execute(select(Plan).where(Plan.id == request.plan_id, Plan.is_active == True))
    plan = plan_result.scalar_one_or_none()
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")

    doc_type = request.doc_type
    if doc_type not in [e.value for e in DocumentType]:
        raise HTTPException(status_code=400, detail="Invalid document type")

    # Build document content
    now = datetime.utcnow()
    doc_title = f"{plan.name or '巡察'}_{doc_type}_{now.strftime('%Y%m%d')}"

    # Generate a simple Excel file as document placeholder
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = doc_type

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1677FF")
    header_align = Alignment(horizontal="center", vertical="center")

    ws.append([doc_type])
    ws.append(["巡察计划名称", plan.name or ""])
    ws.append(["年份", plan.year or ""])
    ws.append(["轮次", plan.round_name or ""])
    ws.append(["巡察范围", plan.scope or ""])
    ws.append(["计划开始日期", plan.planned_start_date.strftime('%Y-%m-%d') if plan.planned_start_date else ""])
    ws.append(["计划结束日期", plan.planned_end_date.strftime('%Y-%m-%d') if plan.planned_end_date else ""])
    ws.append(["生成日期", now.strftime('%Y-%m-%d')])
    ws.append(["生成人", current_user.full_name or current_user.username])

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Save file
    file_name = f"{doc_title}.xlsx"
    file_path = DOCUMENTS_DIR / file_name
    wb.save(file_path)

    # Create document record
    doc = Document(
        title=doc_title,
        doc_number=f"DOC-{now.strftime('%Y%m%d%H%M%S')}",
        type=doc_type,
        generate_date=now,
        generator=current_user.id,
        file_path=str(file_path),
        file_url=f"/documents/{file_name}",
        plan_id=request.plan_id,
    )
    db.add(doc)
    await db.commit()
    await db.refresh(doc)
    await write_audit_log(db, current_user.id, "generate", "document", doc.id, {"type": doc_type, "plan_id": str(request.plan_id)})

    return doc


@router.post("/generate-rectification-notice")
async def generate_rectification_notice(
    request: GenerateRectificationNoticeRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Generate rectification notice document."""
    rect_result = await db.execute(select(Rectification).where(Rectification.id == request.rectification_id, Rectification.is_active == True))
    rect = rect_result.scalar_one_or_none()
    if not rect:
        raise HTTPException(status_code=404, detail="Rectification not found")

    now = datetime.utcnow()
    doc_title = f"整改通知书_{rect.title[:20] if rect.title else '未知'}_{now.strftime('%Y%m%d')}"

    # Generate Excel document
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "整改通知书"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1677FF")
    header_align = Alignment(horizontal="center", vertical="center")

    ws.append(["整改通知书"])
    ws.append(["问题标题", rect.title or ""])
    ws.append(["问题描述", rect.problem_description or ""])
    ws.append(["整改要求", rect.rectification_requirement or ""])
    ws.append(["截止日期", rect.deadline.strftime('%Y-%m-%d') if rect.deadline else ""])
    ws.append(["当前状态", rect.status or ""])
    ws.append(["完成进度", f"{rect.progress or 0}%"])
    ws.append(["生成日期", now.strftime('%Y-%m-%d')])
    ws.append(["生成人", current_user.full_name or current_user.username])

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 6, 60)

    # Save file
    file_name = f"{doc_title}.xlsx"
    file_path = DOCUMENTS_DIR / file_name
    wb.save(file_path)

    # Create document record
    doc = Document(
        title=doc_title,
        doc_number=f"RECT-{now.strftime('%Y%m%d%H%M%S')}",
        type=DocumentType.RECTIFICATION_NOTICE.value,
        generate_date=now,
        generator=current_user.id,
        file_path=str(file_path),
        file_url=f"/documents/{file_name}",
        rectification_id=request.rectification_id,
    )
    db.add(doc)
    await db.commit()
    await db.refresh(doc)
    await write_audit_log(db, current_user.id, "generate", "rectification_notice", doc.id, {"rectification_id": str(request.rectification_id)})

    return doc
