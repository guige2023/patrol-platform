from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from typing import List, Optional
from uuid import UUID
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.dependencies import get_db, get_current_user
from app.models.draft import Draft, DraftAttachment
from app.models.user import User
from app.models.inspection_group import InspectionGroup
from app.models.unit import Unit
from app.schemas.draft import DraftCreate, DraftUpdate, DraftResponse, DraftSubmitRequest
from app.schemas.common import PaginatedResponse, PageResult
from app.core.audit import write_audit_log

router = APIRouter()


@router.get("/", response_model=PaginatedResponse[DraftResponse])
async def list_drafts(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=9999),
    title: Optional[str] = None,
    status: Optional[str] = None,
    group_id: Optional[UUID] = None,
    unit_id: Optional[UUID] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = select(Draft).where(Draft.is_active == True)
    if title:
        query = query.where(Draft.title.ilike(f"%{title}%"))
    if status:
        query = query.where(Draft.status == status)
    if group_id:
        query = query.where(Draft.group_id == group_id)
    if unit_id:
        query = query.where(Draft.unit_id == unit_id)
    
    count_result = await db.execute(select(func.count()).select_from(query.subquery()))
    total = count_result.scalar()
    
    query = query.order_by(Draft.created_at.desc()).offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(query)
    items = result.scalars().all()
    
    return PaginatedResponse(
        data=PageResult(items=items, total=total, page=page, page_size=page_size)
    )


STATUS_LABELS = {
    "draft": "草稿",
    "preliminary_review": "初审",
    "final_review": "终审",
    "approved": "已批准",
    "rejected": "已驳回",
}

SEVERITY_LABELS = {
    "mild": "轻微",
    "moderate": "中等",
    "severe": "严重",
}


@router.get("/download")
async def export_drafts(
    status: Optional[str] = None,
    category: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export all drafts as .xlsx."""
    query = select(Draft).where(Draft.is_active == True)
    if status:
        query = query.where(Draft.status == status)
    if category:
        query = query.where(Draft.category == category)
    query = query.order_by(Draft.created_at.desc()).limit(10000)
    result = await db.execute(query)
    drafts = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "底稿记录"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1677FF")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = [
        "标题", "类别", "问题类型", "严重程度", "证据摘要",
        "状态", "初审意见", "终审意见",
        "创建时间",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for d in drafts:
        created = d.created_at.strftime('%Y-%m-%d %H:%M') if d.created_at else ""
        ws.append([
            d.title or "",
            d.category or "",
            d.problem_type or "",
            SEVERITY_LABELS.get(d.severity, d.severity or ""),
            d.evidence_summary or "",
            STATUS_LABELS.get(d.status, d.status or ""),
            d.preliminary_review_comment or "",
            d.final_review_comment or "",
            created,
        ])

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''drafts.xlsx"},
    )


@router.get("/{draft_id}", response_model=DraftResponse)
async def get_draft(draft_id: UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(Draft).where(Draft.id == draft_id))
    draft = result.scalar_one_or_none()
    if not draft:
        raise HTTPException(status_code=404, detail="Draft not found")

    group_name = None
    if draft.group_id:
        grp_result = await db.execute(select(InspectionGroup).where(InspectionGroup.id == draft.group_id))
        grp = grp_result.scalar_one_or_none()
        if grp:
            group_name = grp.name

    unit_name = None
    if draft.unit_id:
        unit_result = await db.execute(select(Unit).where(Unit.id == draft.unit_id))
        unit = unit_result.scalar_one_or_none()
        if unit:
            unit_name = unit.name

    return {
        **draft.__dict__,
        "group_name": group_name,
        "unit_name": unit_name,
    }


@router.post("/", response_model=DraftResponse, status_code=201)
async def create_draft(draft_data: DraftCreate, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    draft = Draft(**draft_data.model_dump(), created_by=current_user.id)
    db.add(draft)
    await db.commit()
    await db.refresh(draft)
    await write_audit_log(db, current_user.id, "create", "draft", draft.id, {"title": draft.title})
    return draft


@router.put("/{draft_id}", response_model=DraftResponse)
async def update_draft(draft_id: UUID, draft_data: DraftUpdate, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(Draft).where(Draft.id == draft_id))
    draft = result.scalar_one_or_none()
    if not draft:
        raise HTTPException(status_code=404, detail="Draft not found")
    for key, value in draft_data.model_dump(exclude_unset=True).items():
        setattr(draft, key, value)
    await db.commit()
    await db.refresh(draft)
    await write_audit_log(db, current_user.id, "update", "draft", draft.id, {"title": draft.title})
    return draft


@router.post("/{draft_id}/submit")
async def submit_draft_action(draft_id: UUID, request: DraftSubmitRequest, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(Draft).where(Draft.id == draft_id))
    draft = result.scalar_one_or_none()
    if not draft:
        raise HTTPException(status_code=404, detail="Draft not found")
    
    action = request.action
    if action == "submit":
        if draft.status != "draft":
            raise HTTPException(status_code=400, detail="Only draft can be submitted")
        draft.status = "preliminary_review"
    elif action == "preliminary_review":
        if draft.status != "preliminary_review":
            raise HTTPException(status_code=400, detail="Wrong status")
        draft.preliminary_reviewer = current_user.id
        draft.preliminary_review_comment = request.comment
        draft.preliminary_review_at = func.now()
        draft.status = "final_review"
    elif action == "final_review":
        if draft.status != "final_review":
            raise HTTPException(status_code=400, detail="Wrong status")
        draft.final_reviewer = current_user.id
        draft.final_review_comment = request.comment
        draft.final_review_at = func.now()
        draft.status = "approved"
    elif action == "approve":
        draft.approved_by = current_user.id
        draft.approved_at = func.now()
        draft.status = "approved"
    elif action == "reject":
        draft.status = "rejected"
    
    await db.commit()
    await write_audit_log(db, current_user.id, f"draft_{action}", "draft", draft_id, {})
    return {"message": f"Draft {action} success", "status": draft.status}


@router.delete("/{draft_id}")
async def delete_draft(draft_id: UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(Draft).where(Draft.id == draft_id))
    draft = result.scalar_one_or_none()
    if not draft:
        raise HTTPException(status_code=404, detail="Draft not found")
    draft.is_active = False
    await db.commit()
    await write_audit_log(db, current_user.id, "delete", "draft", draft_id, {})
    return {"message": "Draft deleted"}


@router.post("/batch-delete")
async def batch_delete_drafts(
    ids: List[UUID],
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Soft-delete multiple drafts at once."""
    if not ids:
        raise HTTPException(status_code=400, detail="No IDs provided")
    result = await db.execute(
        select(Draft).where(Draft.id.in_(ids), Draft.is_active == True)
    )
    drafts = result.scalars().all()
    if not drafts:
        raise HTTPException(status_code=404, detail="No drafts found")
    for d in drafts:
        d.is_active = False
    await db.commit()
    for d in drafts:
        await write_audit_log(db, current_user.id, "delete", "draft", d.id, {"title": d.title})
    return {"message": f"{len(drafts)} drafts deleted"}
