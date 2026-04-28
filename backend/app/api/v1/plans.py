from fastapi import APIRouter, Depends, HTTPException, Query
from pydantic import BaseModel
from fastapi.responses import Response, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload
from typing import List, Optional
from uuid import UUID
from datetime import datetime
from urllib.parse import quote
from app.dependencies import get_uow, get_current_user
from app.database import UnitOfWork
from app.models.plan import Plan, PlanVersion, PlanStatus
from app.models.user import User
from app.models.inspection_group import InspectionGroup, GroupMember
from app.models.draft import Draft
from app.models.rectification import Rectification
from app.models.cadre import Cadre
from app.models.unit import Unit
from app.schemas.plan import PlanCreate, PlanUpdate, PlanResponse
from app.schemas.common import PaginatedResponse, PageResult
from app.schemas.common import Response as ResponseWrapper
from app.core.audit import write_audit_log
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os

# 注册中文字体（宋体）
FONT_PATH = "/System/Library/Fonts/Supplemental/Songti.ttc"
if os.path.exists(FONT_PATH):
    pdfmetrics.registerFont(TTFont('Songti', FONT_PATH))
    CHINESE_FONT = 'Songti'
else:
    CHINESE_FONT = 'Helvetica'

router = APIRouter()


@router.get("/", response_model=PaginatedResponse[PlanResponse])
async def list_plans(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=9999),
    name: Optional[str] = None,
    year: Optional[int] = None,
    status: Optional[str] = None,
    principal_id: Optional[UUID] = None,
    uow: UnitOfWork = Depends(get_uow),
    current_user: User = Depends(get_current_user),
):
    query = select(Plan).where(Plan.is_active == True)
    if name:
        query = query.where(Plan.name.ilike(f"%{name}%"))
    if year:
        query = query.where(Plan.year == year)
    if status:
        query = query.where(Plan.status == status)
    if principal_id:
        query = query.where(Plan.created_by == principal_id)

    count_result = await uow.execute(select(func.count()).select_from(query.subquery()))
    total = count_result.scalar()

    query = query.order_by(Plan.created_at.desc()).offset((page - 1) * page_size).limit(page_size)
    result = await uow.execute(query)
    items = result.scalars().all()

    # Fetch principal names
    user_ids = list({item.created_by for item in items if item.created_by})
    user_map: dict[str, str] = {}
    if user_ids:
        user_result = await uow.execute(
            select(User.id, User.full_name).where(User.id.in_(user_ids))
        )
        user_map = {str(uid): fname for uid, fname in user_result.all()}

    response_items = [
        {**{"id": item.id, "name": item.name, "year": item.year, "status": item.status,
          "created_by": item.created_by, "created_at": item.created_at,
          "planned_start_date": item.planned_start_date, "planned_end_date": item.planned_end_date,
          "round_name": item.round_name, "is_active": item.is_active,
          "principal_name": user_map.get(str(item.created_by), "")},
         **{k: v for k, v in item.__dict__.items()
            if k not in ("id", "name", "year", "status", "created_by", "created_at",
                         "planned_start_date", "planned_end_date", "round_name", "is_active")}}
        for item in items
    ]

    return PaginatedResponse(
        data=PageResult(items=response_items, total=total, page=page, page_size=page_size)
    )


@router.get("/years")
async def get_plan_years(
    uow: UnitOfWork = Depends(get_uow),
    current_user: User = Depends(get_current_user),
):
    """Return all distinct years that have plans."""
    result = await uow.execute(
        select(Plan.year).where(Plan.is_active == True).distinct().order_by(Plan.year.desc())
    )
    years = [row[0] for row in result.all() if row[0]]
    return {"data": years}


STATUS_LABELS = {
    "draft": "草稿",
    "submitted": "已提交",
    "approved": "已批准",
    "published": "已发布",
    "in_progress": "进行中",
    "completed": "已完成",
}


@router.get("/download")
async def export_plans(
    year: Optional[int] = None,
    status: Optional[str] = None,
    ids: Optional[str] = None,
    uow: UnitOfWork = Depends(get_uow),
    current_user: User = Depends(get_current_user),
):
    """Export plans as .xlsx. If ids is provided (comma-separated UUIDs), export only those."""
    query = select(Plan).where(Plan.is_active == True)
    if year:
        query = query.where(Plan.year == year)
    if status:
        query = query.where(Plan.status == status)
    if ids:
        id_list = [i.strip() for i in ids.split(",") if i.strip()]
        if id_list:
            query = query.where(Plan.id.in_(id_list))
    query = query.order_by(Plan.year.desc(), Plan.created_at.desc()).limit(10000)
    result = await uow.execute(query)
    plans = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "巡察计划"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1677FF")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = [
        "计划名称", "轮次", "年份", "状态",
        "计划开始日期", "计划结束日期", "实际开始日期", "实际结束日期",
        "巡察范围", "重点领域", "版本", "审批意见",
        "授权日期", "创建时间",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for p in plans:
        focus_str = ",".join(p.focus_areas) if isinstance(p.focus_areas, list) else (p.focus_areas or "")
        auth_date = p.authorization_date.strftime('%Y-%m-%d') if p.authorization_date else ""
        created = p.created_at.strftime('%Y-%m-%d %H:%M') if p.created_at else ""
        ws.append([
            p.name or "",
            p.round_name or "",
            p.year or "",
            STATUS_LABELS.get(p.status, p.status or ""),
            p.planned_start_date.strftime('%Y-%m-%d') if p.planned_start_date else "",
            p.planned_end_date.strftime('%Y-%m-%d') if p.planned_end_date else "",
            p.actual_start_date.strftime('%Y-%m-%d') if p.actual_start_date else "",
            p.actual_end_date.strftime('%Y-%m-%d') if p.actual_end_date else "",
            p.scope or "",
            focus_str,
            p.version or 1,
            p.approval_comment or "",
            auth_date,
            created,
        ])

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''plans.xlsx"},
    )


@router.get("/template")
async def download_plan_template(
    current_user: User = Depends(get_current_user),
):
    """Download plan import template (.xlsx)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "巡察计划导入模板"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="52C41A")
    header_align = Alignment(horizontal="center", vertical="center")
    note_fill = PatternFill("solid", fgColor="FFF7E6")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = ["计划名称*", "轮次(第X轮)", "年份*", "计划开始日期", "计划结束日期", "巡察范围", "状态"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    notes = [
        "必填", "如：第一轮", "必填，如：2024", "YYYY-MM-DD", "YYYY-MM-DD", "巡察范围描述", "draft/submitted/approved/published/in_progress/completed",
    ]
    ws.append(notes)
    for cell in ws[2]:
        cell.fill = note_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        cell.font = Font(size=10, color="999999")

    sample = [
        ["2024年巡察计划", "第一轮", 2024, "2024-03-01", "2024-05-31", "对XX单位开展巡察", "draft"],
    ]
    for row in sample:
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[2].height = 36
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 6, 45)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''plan_template.xlsx"},
    )


RECTIFICATION_STATUS_LABELS = {
    "dispatched": "已派发",
    "signed": "已签收",
    "progressing": "进行中",
    "completed": "已完成",
    "submitted": "待验收",
    "verified": "已验收",
    "rejected": "已驳回",
}

DRAFT_STATUS_LABELS = {
    "draft": "草稿",
    "preliminary_review": "初审中",
    "final_review": "终审中",
    "approved": "已通过",
    "rejected": "已驳回",
}

DRAFT_SEVERITY_LABELS = {
    "mild": "轻微",
    "moderate": "中等",
    "severe": "严重",
}


def _make_header(ws, headers):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1677FF")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)


@router.get("/{plan_id}/report")
async def export_plan_report(
    plan_id: UUID,
    uow: UnitOfWork = Depends(get_uow),
    current_user: User = Depends(get_current_user),
):
    """Export complete inspection plan report as multi-sheet Excel.

    Sheets: 巡察计划 | 巡察组 | 整改项 | 底稿
    """
    # 1. Plan
    result = await uow.execute(select(Plan).where(Plan.id == plan_id, Plan.is_active == True))
    plan = result.scalar_one_or_none()
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")

    # 2. Groups for this plan
    groups_result = await uow.execute(
        select(InspectionGroup)
        .options(selectinload(InspectionGroup.members).selectinload(GroupMember.cadre))
        .where(InspectionGroup.plan_id == plan_id, InspectionGroup.is_active == True)
    )
    groups = groups_result.scalars().all()
    group_ids = [g.id for g in groups]
    target_unit_ids = list({g.target_unit_id for g in groups if g.target_unit_id})

    # 3. Drafts for these groups
    drafts = []
    if group_ids:
        drafts_result = await uow.execute(
            select(Draft).where(Draft.group_id.in_(group_ids), Draft.is_active == True)
        )
        drafts = drafts_result.scalars().all()

    # 4. Rectifications for target units
    rectifications = []
    if target_unit_ids:
        rect_result = await uow.execute(
            select(Rectification).where(
                Rectification.unit_id.in_(target_unit_ids),
                Rectification.is_active == True
            )
        )
        rectifications = rect_result.scalars().all()

    # 5. Build workbook
    wb = openpyxl.Workbook()

    # ── Sheet 1: 巡察计划 ──
    ws1 = wb.active
    ws1.title = "巡察计划"
    _make_header(ws1, [
        "计划名称", "年份", "轮次", "巡察范围", "重点领域",
        "授权文书", "授权日期",
        "计划开始日期", "计划结束日期",
        "实际开始日期", "实际结束日期",
        "状态", "审批意见", "版本",
    ])
    focus_str = ",".join(plan.focus_areas) if isinstance(plan.focus_areas, list) else (plan.focus_areas or "")
    ws1.append([
        plan.name or "",
        plan.year or "",
        plan.round_name or "",
        plan.scope or "",
        focus_str,
        plan.authorization_letter or "",
        plan.authorization_date.strftime('%Y-%m-%d') if plan.authorization_date else "",
        plan.planned_start_date.strftime('%Y-%m-%d') if plan.planned_start_date else "",
        plan.planned_end_date.strftime('%Y-%m-%d') if plan.planned_end_date else "",
        plan.actual_start_date.strftime('%Y-%m-%d') if plan.actual_start_date else "",
        plan.actual_end_date.strftime('%Y-%m-%d') if plan.actual_end_date else "",
        STATUS_LABELS.get(plan.status, plan.status or ""),
        plan.approval_comment or "",
        plan.version or 1,
    ])
    _auto_width(ws1)

    # ── Sheet 2: 巡察组 ──
    ws2 = wb.create_sheet("巡察组")
    _make_header(ws2, [
        "巡察组名称", "状态", "目标单位", "组长", "副组长", "联络员",
        "组员", "授权文书", "授权日期",
    ])
    for g in groups:
        member_by_role = {"组长": [], "副组长": [], "联络员": [], "组员": []}
        for m in g.members:
            if m.cadre:
                name = m.cadre.name or "未知"
            else:
                name = "未知"
            role = m.role or "组员"
            if role in member_by_role:
                member_by_role[role].append(name)
            else:
                member_by_role["组员"].append(name)
        target_unit_name = ""
        if g.target_unit_id:
            u_result = await uow.execute(select(Unit).where(Unit.id == g.target_unit_id))
            u = u_result.scalar_one_or_none()
            if u:
                target_unit_name = u.name
        ws2.append([
            g.name or "",
            STATUS_LABELS.get(g.status, g.status or ""),
            target_unit_name,
            "、".join(member_by_role["组长"]),
            "、".join(member_by_role["副组长"]),
            "、".join(member_by_role["联络员"]),
            "、".join(member_by_role["组员"]),
            g.authorization_letter or "",
            g.authorization_date.strftime('%Y-%m-%d') if g.authorization_date else "",
        ])
    _auto_width(ws2)

    # ── Sheet 3: 整改项 ──
    ws3 = wb.create_sheet("整改项")
    _make_header(ws3, [
        "整改标题", "问题描述", "整改要求", "责任单位",
        "状态", "进度%", "派发日期", "签收日期", "完成日期", "整改要求",
    ])
    for r in rectifications:
        unit_name = ""
        if r.unit_id:
            u_result = await uow.execute(select(Unit).where(Unit.id == r.unit_id))
            u = u_result.scalar_one_or_none()
            if u:
                unit_name = u.name
        ws3.append([
            r.title or "",
            r.problem_description or "",
            r.rectification_requirement or "",
            unit_name,
            RECTIFICATION_STATUS_LABELS.get(r.status, r.status or ""),
            r.progress or 0,
            r.created_at.strftime('%Y-%m-%d') if r.created_at else "",
            r.sign_date.strftime('%Y-%m-%d') if r.sign_date else "",
            r.completed_at.strftime('%Y-%m-%d') if r.completed_at else "",
            r.rectification_requirement or "",
        ])
    _auto_width(ws3)

    # ── Sheet 4: 底稿 ──
    ws4 = wb.create_sheet("底稿")
    _make_header(ws4, [
        "底稿标题", "问题类型", "严重程度", "分类",
        "状态", "证据摘要", "初审意见", "终审意见",
    ])
    for d in drafts:
        ws4.append([
            d.title or "",
            d.problem_type or "",
            DRAFT_SEVERITY_LABELS.get(d.severity, d.severity or ""),
            d.category or "",
            DRAFT_STATUS_LABELS.get(d.status, d.status or ""),
            d.evidence_summary or "",
            d.preliminary_review_comment or "",
            d.review_comment or "",
        ])
    _auto_width(ws4)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    plan_name_safe = (plan.name or "巡察报告").replace("/", "_").replace("\\", "_")
    filename = f"{plan_name_safe}_完整报告.xlsx"
    encoded_filename = quote(filename)
    return Response(
        output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"},
    )


@router.get("/{plan_id}/checklist")
async def export_plan_checklist_pdf(
    plan_id: UUID,
    uow: UnitOfWork = Depends(get_uow),
    current_user: User = Depends(get_current_user),
):
    """Export inspection checklist as printable PDF."""
    # 1. Plan
    result = await uow.execute(select(Plan).where(Plan.id == plan_id, Plan.is_active == True))
    plan = result.scalar_one_or_none()
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")

    # 2. Groups with members + unit names
    groups_result = await uow.execute(
        select(InspectionGroup)
        .options(selectinload(InspectionGroup.members).selectinload(GroupMember.cadre))
        .where(InspectionGroup.plan_id == plan_id, InspectionGroup.is_active == True)
    )
    groups = groups_result.scalars().all()

    # Build unit map
    all_unit_ids = list({g.target_unit_id for g in groups if g.target_unit_id})
    unit_map = {}
    if all_unit_ids:
        units_result = await uow.execute(select(Unit).where(Unit.id.in_(all_unit_ids)))
        for u in units_result.scalars().all():
            unit_map[u.id] = u.name

    # 3. Rectifications for target units
    rectifications = []
    if all_unit_ids:
        rect_result = await uow.execute(
            select(Rectification).where(
                Rectification.unit_id.in_(all_unit_ids),
                Rectification.is_active == True
            )
        )
        rectifications = rect_result.scalars().all()

    # 4. Build PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=20*mm, rightMargin=20*mm,
        topMargin=20*mm, bottomMargin=20*mm
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Title'], fontName=CHINESE_FONT, fontSize=16, alignment=TA_CENTER, spaceAfter=6)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontName=CHINESE_FONT, fontSize=10, alignment=TA_CENTER, textColor=colors.grey, spaceAfter=12)
    section_style = ParagraphStyle('Section', parent=styles['Heading2'], fontName=CHINESE_FONT, fontSize=12, spaceBefore=12, spaceAfter=4, textColor=colors.HexColor('#1677ff'))
    normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontName=CHINESE_FONT, fontSize=9)

    story = []

    # ── 封面 ──
    story.append(Paragraph(plan.name or "巡察检查清单", title_style))
    year_str = f"{plan.year}年" if plan.year else ""
    round_str = plan.round_name or ""
    story.append(Paragraph(f"{year_str} {round_str} 巡察检查清单  |  生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#1677ff'), spaceAfter=12))

    # ── 计划基本信息 ──
    story.append(Paragraph("一、计划信息", section_style))
    info_data = [
        ["计划名称", plan.name or "-"],
        ["年　　份", str(plan.year) if plan.year else "-"],
        ["轮　　次", plan.round_name or "-"],
        ["巡察范围", plan.scope or "-"],
        ["重点领域", ", ".join(plan.focus_areas) if plan.focus_areas else "-"],
        ["授权文书", plan.authorization_letter or "-"],
        ["计划起止", f"{plan.planned_start_date or '-'} ~ {plan.planned_end_date or '-'}"],
        ["实际起止", f"{plan.actual_start_date or '-'} ~ {plan.actual_end_date or '-'}"],
    ]
    info_table = Table(info_data, colWidths=[40*mm, 130*mm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), CHINESE_FONT),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('FONTNAME', (0,0), (0,-1), CHINESE_FONT),
        ('FONTNAME', (1,0), (1,-1), CHINESE_FONT),
        ('TEXTCOLOR', (0,0), (0,-1), colors.HexColor('#555555')),
        ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#f5f5f5')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dddddd')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('LEFTPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 8*mm))

    # ── 巡察组信息 ──
    story.append(Paragraph("二、巡察组信息", section_style))
    if groups:
        group_header = ["序号", "巡察组名称", "被巡察单位", "组长", "副组长", "联络员", "组员", "状态"]
        group_rows = []
        for i, g in enumerate(groups, 1):
            unit_name = unit_map.get(g.target_unit_id, "-") if g.target_unit_id else "-"
            members = g.members or []
            leader = next((m.cadre.name if m.cadre else "" for m in members if m.is_leader), "-")
            vice = ", ".join([m.cadre.name if m.cadre else "" for m in members if not m.is_leader and getattr(m, 'role', None) == '副组长']) or "-"
            liaison = ", ".join([m.cadre.name if m.cadre else "" for m in members if not m.is_leader and getattr(m, 'role', None) == '联络员']) or "-"
            members_names = ", ".join([m.cadre.name if m.cadre else "" for m in members if m.cadre]) or "-"
            STATUS_LABELS_GROUP = {"draft": "草稿", "approved": "已审批", "active": "进行中", "completed": "已完成"}
            group_rows.append([
                str(i),
                g.name or "-",
                unit_name,
                leader,
                vice,
                liaison,
                members_names[:30],
                STATUS_LABELS_GROUP.get(g.status, g.status or "-"),
            ])
        group_table = Table([group_header] + group_rows, colWidths=[10*mm, 30*mm, 30*mm, 20*mm, 20*mm, 18*mm, 40*mm, 14*mm])
        group_table.setStyle(TableStyle([
            ('FONTNAME', (0,0), (-1,-1), CHINESE_FONT),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1677ff')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), CHINESE_FONT),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('ALIGN', (0,1), (0,-1), 'CENTER'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cccccc')),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 3),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
            ('LEFTPADDING', (0,0), (-1,-1), 4),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f9f9f9')]),
        ]))
        story.append(group_table)
    else:
        story.append(Paragraph("暂无巡察组数据", normal_style))
    story.append(Spacer(1, 8*mm))

    # ── 整改项检查 ──
    story.append(Paragraph("三、整改项检查", section_style))
    if rectifications:
        rect_header = ["序号", "标题", "责任单位", "状态", "进度", "截止日期", "派发日期"]
        rect_rows = []
        RECT_STATUS = {"dispatched": "已派发", "signed": "已签收", "progressing": "整改中", "completed": "已完成", "submitted": "待验收", "verified": "已验收", "rejected": "已驳回"}
        for i, r in enumerate(rectifications[:50], 1):  # limit to 50
            unit_name = unit_map.get(r.unit_id, "-") if r.unit_id else "-"
            rect_rows.append([
                str(i),
                (r.title or "-")[:35],
                unit_name[:15],
                RECT_STATUS.get(r.status, r.status or "-"),
                f"{r.progress or 0}%",
                r.deadline or "-",
                r.created_at.strftime('%Y-%m-%d') if r.created_at else "-",
            ])
        rect_table = Table([rect_header] + rect_rows, colWidths=[10*mm, 55*mm, 30*mm, 20*mm, 15*mm, 25*mm, 17*mm])
        rect_table.setStyle(TableStyle([
            ('FONTNAME', (0,0), (-1,-1), CHINESE_FONT),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#52c41a')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), CHINESE_FONT),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('ALIGN', (0,1), (0,-1), 'CENTER'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cccccc')),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 3),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
            ('LEFTPADDING', (0,0), (-1,-1), 4),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f9f9f9')]),
        ]))
        story.append(rect_table)
    else:
        story.append(Paragraph("暂无整改项数据", normal_style))
    story.append(Spacer(1, 8*mm))

    # ── 签字栏 ──
    story.append(Paragraph("四、签字确认", section_style))
    sign_data = [
        ["巡察组组长签字：", "签字日期："],
        ["被巡察单位负责人签字：", "签字日期："],
        ["巡察办审核签字：", "签字日期："],
    ]
    sign_table = Table(sign_data, colWidths=[85*mm, 85*mm])
    sign_table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), CHINESE_FONT),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('TOPPADDING', (0,0), (-1,-1), 14),
        ('BOTTOMPADDING', (0,0), (-1,-1), 14),
        ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.HexColor('#cccccc')),
    ]))
    story.append(sign_table)

    doc.build(story)
    buffer.seek(0)
    plan_name_safe = (plan.name or "巡察检查清单").replace("/", "_").replace("\\", "_")
    filename = f"{plan_name_safe}_检查清单.pdf"
    encoded_filename = quote(filename)
    return Response(
        buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"},
    )


@router.get("/{plan_id}/feedback")
async def get_plan_feedback(
    plan_id: UUID,
    uow: UnitOfWork = Depends(get_uow),
    current_user: User = Depends(get_current_user),
):
    """Get all feedback for a plan: rectification confirmations + draft reviews."""
    from app.models.draft import Draft
    from app.models.rectification import Rectification

    # Verify plan exists
    plan_result = await uow.execute(select(Plan).where(Plan.id == plan_id, Plan.is_active == True))
    plan = plan_result.scalar_one_or_none()
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")

    # Get groups for this plan
    groups_result = await uow.execute(
        select(InspectionGroup).where(InspectionGroup.plan_id == plan_id, InspectionGroup.is_active == True)
    )
    groups = groups_result.scalars().all()
    group_ids = [g.id for g in groups]
    target_unit_ids = list({g.target_unit_id for g in groups if g.target_unit_id})

    # Get confirmed rectifications (feedback from units)
    confirmed_rects = []
    if target_unit_ids:
        rect_result = await uow.execute(
            select(Rectification).where(
                Rectification.unit_id.in_(target_unit_ids),
                Rectification.is_active == True,
                Rectification.confirmed_completed == True,
            )
        )
        rectifications = rect_result.scalars().all()
        for r in rectifications:
            unit_result = await uow.execute(select(Unit).where(Unit.id == r.unit_id))
            unit = unit_result.scalar_one_or_none()
            confirmed_rects.append({
                "rectification_id": str(r.id),
                "unit_id": str(r.unit_id) if r.unit_id else None,
                "unit_name": unit.name if unit else None,
                "problem_description": r.problem_description,
                "confirmed_at": r.confirmed_at.isoformat() if r.confirmed_at else None,
                "confirm_notes": r.confirm_notes,
            })

    # Get draft reviews (feedback from supervisors)
    draft_feedback = []
    if group_ids:
        draft_result = await uow.execute(
            select(Draft).where(Draft.group_id.in_(group_ids), Draft.is_active == True)
        )
        drafts = draft_result.scalars().all()
        for d in drafts:
            if d.preliminary_review_comment or d.final_review_comment:
                draft_feedback.append({
                    "draft_id": str(d.id),
                    "title": d.title,
                    "category": d.category,
                    "status": d.status,
                    "preliminary_review_comment": d.preliminary_review_comment,
                    "final_review_comment": d.final_review_comment,
                })

    return {
        "plan_id": str(plan_id),
        "plan_name": plan.name,
        "confirmed_rectifications": confirmed_rects,
        "draft_feedback": draft_feedback,
        "summary": {
            "total_confirmed_rectifications": len(confirmed_rects),
            "total_draft_feedback": len(draft_feedback),
        },
    }


@router.get("/{plan_id}/cadre-export")
async def export_plan_cadres(
    plan_id: UUID,
    uow: UnitOfWork = Depends(get_uow),
    current_user: User = Depends(get_current_user),
):
    """Export all cadres (inspectors) for a plan as Excel."""
    # Verify plan exists
    plan_result = await uow.execute(select(Plan).where(Plan.id == plan_id, Plan.is_active == True))
    plan = plan_result.scalar_one_or_none()
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")

    # Get groups + members + cadres
    groups_result = await uow.execute(
        select(InspectionGroup)
        .options(selectinload(InspectionGroup.members).selectinload(GroupMember.cadre))
        .where(InspectionGroup.plan_id == plan_id, InspectionGroup.is_active == True)
    )
    groups = groups_result.scalars().all()

    rows = []
    for group in groups:
        for member in (group.members or []):
            cadre = member.cadre
            if not cadre:
                continue
            unit_result = await uow.execute(select(Unit).where(Unit.id == cadre.unit_id))
            unit = unit_result.scalar_one_or_none()
            rows.append([
                group.name or "",
                member.role or "",
                "是" if member.is_leader else "否",
                cadre.name or "",
                cadre.gender or "",
                cadre.position or "",
                cadre.rank or "",
                unit.name if unit else "",
                cadre.phone or "",
                cadre.political_status or "",
            ])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "巡察干部"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1677FF")
    headers = ["巡察组", "角色", "组长", "姓名", "性别", "职务", "职级", "所在单位", "联系电话", "政治面貌"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for row in rows:
        ws.append(row)

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    plan_name_safe = (plan.name or "巡察干部").replace("/", "_").replace("\\", "_")
    filename = f"{plan_name_safe}_干部名单.xlsx"
    encoded_filename = quote(filename)
    return Response(
        output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"},
    )


@router.get("/{plan_id}")
async def get_plan(plan_id: UUID, uow: UnitOfWork = Depends(get_uow), current_user: User = Depends(get_current_user)):
    result = await uow.execute(select(Plan).where(Plan.id == plan_id, Plan.is_active == True))
    plan = result.scalar_one_or_none()
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")
    return {"data": plan, "message": "success"}


@router.post("/", status_code=201)
async def create_plan(plan_data: PlanCreate, uow: UnitOfWork = Depends(get_uow), current_user: User = Depends(get_current_user)):
    plan = Plan(**plan_data.model_dump(), created_by=current_user.id)
    uow.add(plan)
    await uow.commit()
    await uow.refresh(plan)
    await write_audit_log(uow.session, current_user.id, "create", "plan", plan.id, {"name": plan.name})
    
    # Serialize manually for consistent {data, message} format
    plan_dict = {
        "id": str(plan.id),
        "name": plan.name,
        "round_name": plan.round_name,
        "round_number": plan.round_number,
        "year": plan.year,
        "planned_start_date": plan.planned_start_date.isoformat() if plan.planned_start_date else None,
        "planned_end_date": plan.planned_end_date.isoformat() if plan.planned_end_date else None,
        "actual_start_date": plan.actual_start_date.isoformat() if plan.actual_start_date else None,
        "actual_end_date": plan.actual_end_date.isoformat() if plan.actual_end_date else None,
        "scope": plan.scope,
        "focus_areas": plan.focus_areas or [],
        "target_units": plan.target_units or [],
        "authorization_letter": plan.authorization_letter,
        "authorization_date": plan.authorization_date.isoformat() if plan.authorization_date else None,
        "status": plan.status,
        "version": plan.version,
        "version_history": plan.version_history or [],
        "approval_comment": plan.approval_comment,
        "approved_by": str(plan.approved_by) if plan.approved_by else None,
        "is_active": plan.is_active,
        "created_by": str(plan.created_by) if plan.created_by else None,
        "created_at": plan.created_at.isoformat() if plan.created_at else None,
        "updated_at": plan.updated_at.isoformat() if plan.updated_at else None,
    }
    return {"data": plan_dict, "message": "success"}


@router.put("/{plan_id}")
async def update_plan(plan_id: UUID, plan_data: PlanUpdate, uow: UnitOfWork = Depends(get_uow), current_user: User = Depends(get_current_user)):
    result = await uow.execute(select(Plan).where(Plan.id == plan_id))
    plan = result.scalar_one_or_none()
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")
    
    data = plan_data.model_dump(exclude_unset=True)
    if "status" in data and data["status"] != plan.status:
        version = plan.version_history or []
        version.append({"version": plan.version, "date": str(plan.updated_at), "change": f"Status: {plan.status} -> {data['status']}"})
        data["version_history"] = version
    
    for key, value in data.items():
        setattr(plan, key, value)
    
    await uow.commit()
    await uow.refresh(plan)
    await write_audit_log(uow.session, current_user.id, "update", "plan", plan.id, {"name": plan.name})
    
    # Serialize manually for consistent {data, message} format
    plan_dict = {
        "id": str(plan.id),
        "name": plan.name,
        "round_name": plan.round_name,
        "round_number": plan.round_number,
        "year": plan.year,
        "planned_start_date": plan.planned_start_date.isoformat() if plan.planned_start_date else None,
        "planned_end_date": plan.planned_end_date.isoformat() if plan.planned_end_date else None,
        "actual_start_date": plan.actual_start_date.isoformat() if plan.actual_start_date else None,
        "actual_end_date": plan.actual_end_date.isoformat() if plan.actual_end_date else None,
        "scope": plan.scope,
        "focus_areas": plan.focus_areas or [],
        "target_units": plan.target_units or [],
        "authorization_letter": plan.authorization_letter,
        "authorization_date": plan.authorization_date.isoformat() if plan.authorization_date else None,
        "status": plan.status,
        "version": plan.version,
        "version_history": plan.version_history or [],
        "approval_comment": plan.approval_comment,
        "approved_by": str(plan.approved_by) if plan.approved_by else None,
        "is_active": plan.is_active,
        "created_by": str(plan.created_by) if plan.created_by else None,
        "created_at": plan.created_at.isoformat() if plan.created_at else None,
        "updated_at": plan.updated_at.isoformat() if plan.updated_at else None,
    }
    return {"data": plan_dict, "message": "success"}


@router.post("/{plan_id}/submit")
async def submit_plan(plan_id: UUID, uow: UnitOfWork = Depends(get_uow), current_user: User = Depends(get_current_user)):
    result = await uow.execute(select(Plan).where(Plan.id == plan_id))
    plan = result.scalar_one_or_none()
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")
    if plan.status != PlanStatus.DRAFT.value:
        raise HTTPException(status_code=400, detail="Only draft plans can be submitted")
    plan.status = PlanStatus.SUBMITTED.value
    await uow.commit()
    await write_audit_log(uow.session, current_user.id, "submit", "plan", plan.id, {})
    return {"message": "Plan submitted for approval"}


@router.post("/{plan_id}/approve")
async def approve_plan(plan_id: UUID, comment: Optional[str] = None, uow: UnitOfWork = Depends(get_uow), current_user: User = Depends(get_current_user)):
    result = await uow.execute(select(Plan).where(Plan.id == plan_id))
    plan = result.scalar_one_or_none()
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")
    if plan.status != PlanStatus.SUBMITTED.value:
        raise HTTPException(status_code=400, detail="Only submitted plans can be approved")
    plan.status = PlanStatus.APPROVED.value
    plan.approved_by = current_user.id
    plan.approval_comment = comment
    await uow.commit()
    await write_audit_log(uow.session, current_user.id, "approve", "plan", plan.id, {})
    return {"message": "Plan approved"}


@router.post("/{plan_id}/publish")
async def publish_plan(plan_id: UUID, uow: UnitOfWork = Depends(get_uow), current_user: User = Depends(get_current_user)):
    from datetime import datetime
    from app.models.unit import Unit
    result = await uow.execute(select(Plan).where(Plan.id == plan_id))
    plan = result.scalar_one_or_none()
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")
    if plan.status != PlanStatus.APPROVED.value:
        raise HTTPException(status_code=400, detail="Only approved plans can be published")
    plan.status = PlanStatus.PUBLISHED.value

    # 自动更新被巡察单位的巡察年份和历史
    current_year = datetime.now().year
    updated_units = []
    for unit_id in (plan.target_units or []):
        unit_result = await uow.execute(select(Unit).where(Unit.id == unit_id))
        unit = unit_result.scalar_one_or_none()
        if unit:
            unit.last_inspection_year = current_year
            existing = unit.inspection_history or ""
            if plan.round_number is not None:
                round_num_str = f"{plan.round_number}轮"
            else:
                round_num_str = ""
            round_info = f"{current_year}年{plan.round_name or round_num_str}"
            unit.inspection_history = (existing + f"; {round_info}").strip("; ")
            updated_units.append(unit.name)

    await uow.commit()
    await write_audit_log(uow.session, current_user.id, "publish", "plan", plan.id, {
        "target_units": plan.target_units,
        "updated_units": updated_units,
    })
    return {"message": f"Plan published，{len(updated_units)}个单位巡察记录已更新"}


@router.delete("/{plan_id}")
async def delete_plan(plan_id: UUID, uow: UnitOfWork = Depends(get_uow), current_user: User = Depends(get_current_user)):
    result = await uow.execute(select(Plan).where(Plan.id == plan_id))
    plan = result.scalar_one_or_none()
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")
    plan.is_active = False
    await uow.commit()
    await write_audit_log(uow.session, current_user.id, "delete", "plan", plan.id, {"name": plan.name})


# p-flow-1: 通用状态切换（用于 published→in_progress, in_progress→completed）
class StatusUpdateRequest(BaseModel):
    status: str  # "in_progress" | "completed"


@router.post("/{plan_id}/status")
async def update_plan_status(
    plan_id: UUID,
    data: StatusUpdateRequest,
    uow: UnitOfWork = Depends(get_uow),
    current_user: User = Depends(get_current_user),
):
    """切换计划状态：published → in_progress，或 in_progress → completed"""
    result = await uow.execute(select(Plan).where(Plan.id == plan_id))
    plan = result.scalar_one_or_none()
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")

    valid_transitions = {
        PlanStatus.PUBLISHED.value: PlanStatus.IN_PROGRESS.value,
        PlanStatus.IN_PROGRESS.value: PlanStatus.COMPLETED.value,
    }
    if data.status not in valid_transitions.values():
        raise HTTPException(status_code=400, detail="无效的目标状态")

    current_allowed = [k for k, v in valid_transitions.items() if v == data.status]
    if plan.status not in current_allowed:
        raise HTTPException(status_code=400, detail=f"当前状态({plan.status})不可切换至{data.status}")

    if data.status == PlanStatus.IN_PROGRESS.value and plan.status == PlanStatus.PUBLISHED.value:
        plan.status = PlanStatus.IN_PROGRESS.value
        from datetime import datetime
        plan.actual_start_date = datetime.now()
    elif data.status == PlanStatus.COMPLETED.value and plan.status == PlanStatus.IN_PROGRESS.value:
        plan.status = PlanStatus.COMPLETED.value
        from datetime import datetime
        plan.actual_end_date = datetime.now()
    else:
        raise HTTPException(status_code=400, detail="状态流转不符合规则")

    await uow.commit()
    await write_audit_log(uow.session, current_user.id, "status_change", "plan", plan_id, {"new_status": data.status})
    return {"message": f"状态已更新为 {data.status}"}
