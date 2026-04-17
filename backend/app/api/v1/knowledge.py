from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from typing import List, Optional
from uuid import UUID
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.dependencies import get_db, get_current_user
from app.models.knowledge import Knowledge
from app.models.user import User
from app.schemas.knowledge import KnowledgeCreate, KnowledgeUpdate, KnowledgeResponse
from app.schemas.common import PaginatedResponse, PageResult
from app.core.audit import write_audit_log

router = APIRouter()


@router.get("/", response_model=PaginatedResponse[KnowledgeResponse])
async def list_knowledge(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=9999),
    title: Optional[str] = None,
    category: Optional[str] = None,
    tags: Optional[str] = None,
    is_published: Optional[bool] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = select(Knowledge).where(Knowledge.is_active == True)
    if title:
        query = query.where(Knowledge.title.ilike(f"%{title}%"))
    if category:
        query = query.where(Knowledge.category == category)
    if is_published is not None:
        query = query.where(Knowledge.is_published == is_published)
    
    count_result = await db.execute(select(func.count()).select_from(query.subquery()))
    total = count_result.scalar()
    
    query = query.order_by(Knowledge.created_at.desc()).offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(query)
    items = result.scalars().all()
    
    return PaginatedResponse(
        data=PageResult(items=items, total=total, page=page, page_size=page_size)
    )


@router.get("/categories")
async def list_categories(db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(
        select(Knowledge.category, func.count(Knowledge.id).label("count"))
        .where(Knowledge.is_active == True)
        .group_by(Knowledge.category)
    )
    return [{"category": r.category, "count": r.count} for r in result.all()]


CATEGORY_LABELS = {
    "regulation": "制度规定",
    "policy": "政策文件",
    "dict": "名词解释",
}


@router.get("/download")
async def export_knowledge(
    category: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export all knowledge items as .xlsx."""
    query = select(Knowledge).where(Knowledge.is_active == True)
    if category:
        query = query.where(Knowledge.category == category)
    query = query.order_by(Knowledge.created_at.desc()).limit(10000)
    result = await db.execute(query)
    items = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "知识库"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1677FF")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = ["标题", "类别", "版本", "来源", "标签", "是否发布", "创建时间"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for k in items:
        tags_str = ",".join(k.tags) if isinstance(k.tags, list) else (k.tags or "")
        created = k.created_at.strftime('%Y-%m-%d %H:%M') if k.created_at else ""
        ws.append([
            k.title or "",
            CATEGORY_LABELS.get(k.category, k.category or ""),
            k.version or "",
            k.source or "",
            tags_str,
            "是" if k.is_published else "否",
            created,
        ])

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''knowledge.xlsx"},
    )


@router.get("/{knowledge_id}")
async def get_knowledge(knowledge_id: UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(Knowledge).where(Knowledge.id == knowledge_id))
    knowledge = result.scalar_one_or_none()
    if not knowledge:
        raise HTTPException(status_code=404, detail="Knowledge not found")
    return {"data": KnowledgeResponse.model_validate(knowledge), "message": "success"}


@router.post("/", response_model=KnowledgeResponse, status_code=201)
async def create_knowledge(knowledge_data: KnowledgeCreate, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    knowledge = Knowledge(**knowledge_data.model_dump(), created_by=current_user.id)
    db.add(knowledge)
    await db.commit()
    await db.refresh(knowledge)
    await write_audit_log(db, current_user.id, "create", "knowledge", knowledge.id, {"title": knowledge.title})
    return knowledge


@router.put("/{knowledge_id}", response_model=KnowledgeResponse)
async def update_knowledge(knowledge_id: UUID, knowledge_data: KnowledgeUpdate, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(Knowledge).where(Knowledge.id == knowledge_id))
    knowledge = result.scalar_one_or_none()
    if not knowledge:
        raise HTTPException(status_code=404, detail="Knowledge not found")
    
    data = knowledge_data.model_dump(exclude_unset=True)
    if "version" in data and data["version"] != knowledge.version:
        version_history = knowledge.version_history or []
        version_history.append({"version": knowledge.version, "date": str(knowledge.updated_at), "change": "Updated"})
        data["version_history"] = version_history
    
    for key, value in data.items():
        setattr(knowledge, key, value)
    
    await db.commit()
    await db.refresh(knowledge)
    await write_audit_log(db, current_user.id, "update", "knowledge", knowledge.id, {"title": knowledge.title})
    return knowledge


@router.delete("/{knowledge_id}")
async def delete_knowledge(knowledge_id: UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(Knowledge).where(Knowledge.id == knowledge_id))
    knowledge = result.scalar_one_or_none()
    if not knowledge:
        raise HTTPException(status_code=404, detail="Knowledge not found")
    knowledge.is_active = False
    await db.commit()
    await write_audit_log(db, current_user.id, "delete", "knowledge", knowledge.id, {"title": knowledge.title})
    return {"message": "Knowledge deleted"}


@router.post("/{knowledge_id}/publish")
async def publish_knowledge(knowledge_id: UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(Knowledge).where(Knowledge.id == knowledge_id))
    knowledge = result.scalar_one_or_none()
    if not knowledge:
        raise HTTPException(status_code=404, detail="Knowledge not found")
    knowledge.is_published = True
    await db.commit()
    await write_audit_log(db, current_user.id, "publish", "knowledge", knowledge.id, {"title": knowledge.title})
    return {"message": "Knowledge published"}
