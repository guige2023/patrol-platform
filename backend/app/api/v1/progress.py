from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from pydantic import BaseModel
from fastapi.responses import Response, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from typing import Optional
from uuid import UUID
from datetime import datetime
from urllib.parse import quote
from app.dependencies import get_db, get_current_user
from app.models.progress import Progress
from app.models.plan import Plan
from app.models.inspection_group import InspectionGroup
from app.models.user import User
from app.schemas.progress import ProgressCreate, ProgressUpdate, ProgressResponse, GroupOverview
from app.schemas.common import PaginatedResponse, PageResult
from app.core.audit import write_audit_log
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

router = APIRouter()


@router.get("/", response_model=PaginatedResponse[ProgressResponse])
async def list_progress(
    plan_id: Optional[UUID] = None,
    group_id: Optional[UUID] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=9999),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = select(Progress).where(Progress.is_active == True)
    if plan_id:
        query = query.where(Progress.plan_id == plan_id)
    if group_id:
        query = query.where(Progress.group_id == group_id)

    count_result = await db.execute(select(func.count()).select_from(query.subquery()))
    total = count_result.scalar()

    query = query.order_by(Progress.report_date.desc()).offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(query)
    items = result.scalars().all()

    return PaginatedResponse(
        data=PageResult(items=[ProgressResponse.model_validate(item) for item in items], total=total, page=page, page_size=page_size)
    )


@router.get("/group-overview", response_model=list[GroupOverview])
async def get_group_overview(
    plan_id: Optional[UUID] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get overview of progress by group for dashboard cards."""
    query = select(
        Progress.group_id,
        func.count(Progress.id).label("total_reports"),
        func.sum(Progress.talk_count).label("total_talks"),
        func.sum(Progress.doc_review_count).label("total_doc_reviews"),
        func.sum(Progress.petition_count).label("total_petitions"),
        func.sum(Progress.visit_count).label("total_visits"),
        func.sum(Progress.problem_total).label("total_problems"),
        func.max(Progress.report_date).label("latest_report_date"),
    ).where(Progress.is_active == True)

    if plan_id:
        query = query.where(Progress.plan_id == plan_id)

    query = query.group_by(Progress.group_id, Progress.plan_id)
    result = await db.execute(query)
    rows = result.all()

    overviews = []
    for row in rows:
        group_name = None
        plan_name = None
        if row.group_id:
            g_result = await db.execute(select(InspectionGroup.name).where(InspectionGroup.id == row.group_id))
            g = g_result.scalar_one_or_none()
            if g:
                group_name = g

        if plan_id:
            p_result = await db.execute(select(Plan.name).where(Plan.id == plan_id))
        elif row._mapping.get("plan_id"):
            p_result = await db.execute(select(Plan.name).where(Plan.id == row._mapping.get("plan_id")))
        else:
            p_result = None

        if p_result:
            p = p_result.scalar_one_or_none()
            if p:
                plan_name = p

        overviews.append(GroupOverview(
            group_id=row.group_id,
            group_name=group_name,
            plan_id=plan_id or row._mapping.get("plan_id"),
            plan_name=plan_name,
            total_reports=row.total_reports or 0,
            total_talks=int(row.total_talks or 0),
            total_doc_reviews=int(row.total_doc_reviews or 0),
            total_petitions=int(row.total_petitions or 0),
            total_visits=int(row.total_visits or 0),
            total_problems=int(row.total_problems or 0),
            latest_report_date=row.latest_report_date,
        ))

    return overviews


@router.get("/template")
async def download_progress_template(
    current_user: User = Depends(get_current_user),
):
    """Download progress import template (.xlsx)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "巡察进度导入模板"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="52C41A")
    header_align = Alignment(horizontal="center", vertical="center")
    note_fill = PatternFill("solid", fgColor="FFF7E6")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = ["周序号*", "报告日期*", "巡察组ID", "谈话人数", "查阅文档数", "信访数量", "走访数量",
               "问题总数", "党的领导问题数", "党的建设问题数", "重点领域问题数", "下周工作计划", "备注"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    notes = ["必填，数字如：1", "必填，YYYY-MM-DD", "巡察组UUID", "整数", "整数", "整数", "整数",
             "整数", "整数", "整数", "整数", "文本", "文本"]
    ws.append(notes)
    for cell in ws[2]:
        cell.fill = note_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        cell.font = Font(size=10, color="999999")

    ws.row_dimensions[2].height = 36
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 6, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''progress_template.xlsx"},
    )


@router.get("/export")
async def export_progress(
    plan_id: Optional[UUID] = None,
    group_id: Optional[UUID] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export progress data as .xlsx."""
    query = select(Progress).where(Progress.is_active == True)
    if plan_id:
        query = query.where(Progress.plan_id == plan_id)
    if group_id:
        query = query.where(Progress.group_id == group_id)
    query = query.order_by(Progress.report_date.desc()).limit(10000)
    result = await db.execute(query)
    items = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "巡察进度"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1677FF")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = ["周序号", "报告日期", "巡察组", "谈话人数", "查阅文档数", "信访数量", "走访数量",
               "问题总数", "党的领导", "党的建设", "重点领域", "下周计划", "备注"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for p in items:
        group_name = ""
        if p.group_id:
            g_result = await db.execute(select(InspectionGroup.name).where(InspectionGroup.id == p.group_id))
            g = g_result.scalar_one_or_none()
            if g:
                group_name = g
        ws.append([
            p.week_number or "",
            p.report_date.strftime('%Y-%m-%d') if p.report_date else "",
            group_name,
            p.talk_count or 0,
            p.doc_review_count or 0,
            p.petition_count or 0,
            p.visit_count or 0,
            p.problem_total or 0,
            p.problem_party or 0,
            p.problem_pty or 0,
            p.problem_key or 0,
            p.next_week_plan or "",
            p.notes or "",
        ])

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''progress_export.xlsx"},
    )


@router.get("/{progress_id}", response_model=ProgressResponse)
async def get_progress(progress_id: UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(Progress).where(Progress.id == progress_id, Progress.is_active == True))
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="Progress not found")
    return item


@router.post("/", response_model=ProgressResponse, status_code=201)
async def create_progress(progress_data: ProgressCreate, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    progress = Progress(**progress_data.model_dump(), created_by=current_user.id)
    db.add(progress)
    await db.commit()
    await db.refresh(progress)
    await write_audit_log(db, current_user.id, "create", "progress", progress.id, {"week_number": progress.week_number})
    return progress


@router.put("/{progress_id}", response_model=ProgressResponse)
async def update_progress(progress_id: UUID, progress_data: ProgressUpdate, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(Progress).where(Progress.id == progress_id))
    progress = result.scalar_one_or_none()
    if not progress:
        raise HTTPException(status_code=404, detail="Progress not found")

    data = progress_data.model_dump(exclude_unset=True)
    for key, value in data.items():
        setattr(progress, key, value)

    await db.commit()
    await db.refresh(progress)
    await write_audit_log(db, current_user.id, "update", "progress", progress.id, data)
    return progress


@router.delete("/{progress_id}")
async def delete_progress(progress_id: UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(Progress).where(Progress.id == progress_id))
    progress = result.scalar_one_or_none()
    if not progress:
        raise HTTPException(status_code=404, detail="Progress not found")

    progress.is_active = False
    await db.commit()
    await write_audit_log(db, current_user.id, "delete", "progress", progress_id, {})
    return {"message": "deleted"}


@router.post("/import/{plan_id}")
async def import_progress(
    plan_id: UUID,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Import progress from Excel file."""
    plan_result = await db.execute(select(Plan).where(Plan.id == plan_id, Plan.is_active == True))
    plan = plan_result.scalar_one_or_none()
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active

    imported = 0
    errors = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:
            continue
        try:
            week_number = int(row[0])
            report_date_str = row[1]
            if isinstance(report_date_str, datetime):
                report_date = report_date_str
            else:
                report_date = datetime.strptime(str(report_date_str), "%Y-%m-%d")
            group_id = row[2] if row[2] else None
            talk_count = int(row[3]) if row[3] else 0
            doc_review_count = int(row[4]) if row[4] else 0
            petition_count = int(row[5]) if row[5] else 0
            visit_count = int(row[6]) if row[6] else 0
            problem_total = int(row[7]) if row[7] else 0
            problem_party = int(row[8]) if row[8] else 0
            problem_pty = int(row[9]) if row[9] else 0
            problem_key = int(row[10]) if row[10] else 0
            next_week_plan = str(row[11]) if row[11] else None
            notes = str(row[12]) if row[12] else None

            progress = Progress(
                plan_id=plan_id,
                group_id=group_id,
                week_number=week_number,
                report_date=report_date,
                talk_count=talk_count,
                doc_review_count=doc_review_count,
                petition_count=petition_count,
                visit_count=visit_count,
                problem_total=problem_total,
                problem_party=problem_party,
                problem_pty=problem_pty,
                problem_key=problem_key,
                next_week_plan=next_week_plan,
                notes=notes,
                created_by=current_user.id,
            )
            db.add(progress)
            imported += 1
        except Exception as e:
            errors.append(f"Row {row_num}: {str(e)}")

    await db.commit()
    await write_audit_log(db, current_user.id, "import", "progress", plan_id, {"imported": imported})

    return {"imported": imported, "errors": errors}
