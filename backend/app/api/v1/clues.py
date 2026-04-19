from fastapi import APIRouter, Depends, HTTPException, Query, Body
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from uuid import UUID
from typing import Optional, List
from datetime import date, datetime
from app.dependencies import get_uow, get_current_user
from app.database import UnitOfWork
from app.models.clue import Clue
from app.models.user import User
from app.schemas.clue import ClueCreate, ClueUpdate, ClueTransfer, ClueResponse
from app.schemas.common import PaginatedResponse, PageResult
from app.core.audit import write_audit_log
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

router = APIRouter()


@router.get("/", response_model=PaginatedResponse[ClueResponse])
async def list_clues(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=9999),
    title: Optional[str] = None,
    status: Optional[str] = None,
    source: Optional[str] = None,
    category: Optional[str] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    uow: UnitOfWork = Depends(get_uow),
    current_user: User = Depends(get_current_user),
):
    query = select(Clue)
    if title:
        query = query.where(Clue.title.ilike(f"%{title}%"))
    if status:
        query = query.where(Clue.status == status)
    if source:
        query = query.where(Clue.source == source)
    if category:
        query = query.where(Clue.category == category)
    if start_date:
        query = query.where(Clue.created_at >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        query = query.where(Clue.created_at <= datetime.combine(end_date, datetime.max.time()))

    count_result = await uow.execute(select(func.count()).select_from(query.subquery()))
    total = count_result.scalar()

    query = query.order_by(Clue.created_at.desc()).offset((page - 1) * page_size).limit(page_size)
    result = await uow.execute(query)
    items = result.scalars().all()

    return PaginatedResponse(
        data=PageResult(items=items, total=total, page=page, page_size=page_size)
    )


STATUS_LABELS = {
    "registered": "已登记",
    "transferring": "移交中",
    "transferred": "已移交",
    "closed": "已关闭",
}

SEVERITY_LABELS = {
    "low": "一般",
    "medium": "较重",
    "high": "重要",
    "critical": "重大",
}


@router.get("/download")
async def export_clues(
    status: Optional[str] = None,
    source: Optional[str] = None,
    category: Optional[str] = None,
    uow: UnitOfWork = Depends(get_uow),
    current_user: User = Depends(get_current_user),
):
    """Export all clues as .xlsx."""
    query = select(Clue)
    if status:
        query = query.where(Clue.status == status)
    if source:
        query = query.where(Clue.source == source)
    if category:
        query = query.where(Clue.category == category)
    query = query.order_by(Clue.created_at.desc()).limit(10000)
    result = await uow.execute(query)
    clues = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "线索管理"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1677FF")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = [
        "标题", "来源", "类别", "严重程度", "状态",
        "移交目标", "移交时间", "移交备注", "处理结果",
        "创建时间",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for c in clues:
        transfer_date = c.transfer_date.strftime('%Y-%m-%d') if c.transfer_date else ""
        created = c.created_at.strftime('%Y-%m-%d %H:%M') if c.created_at else ""
        ws.append([
            c.title or "",
            c.source or "",
            c.category or "",
            SEVERITY_LABELS.get(c.severity, c.severity or ""),
            STATUS_LABELS.get(c.status, c.status or ""),
            c.transfer_target or "",
            transfer_date,
            c.transfer_comment or "",
            c.handling_result or "",
            created,
        ])

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''clues.xlsx"},
    )


@router.get("/{clue_id}")
async def get_clue(clue_id: UUID, uow: UnitOfWork = Depends(get_uow), current_user: User = Depends(get_current_user)):
    result = await uow.execute(select(Clue).where(Clue.id == clue_id))
    clue = result.scalar_one_or_none()
    if not clue:
        raise HTTPException(status_code=404, detail="Clue not found")
    return {"data": ClueResponse.model_validate(clue), "message": "success"}


@router.post("/", response_model=ClueResponse, status_code=201)
async def create_clue(clue_data: ClueCreate, uow: UnitOfWork = Depends(get_uow), current_user: User = Depends(get_current_user)):
    clue = Clue(**clue_data.model_dump(), registered_by=current_user.id)
    uow.add(clue)
    await uow.commit()
    await uow.refresh(clue)
    await write_audit_log(uow.session, current_user.id, "create", "clue", clue.id, {"title": clue.title})
    return clue


@router.put("/{clue_id}", response_model=ClueResponse)
async def update_clue(clue_id: UUID, clue_data: ClueUpdate, uow: UnitOfWork = Depends(get_uow), current_user: User = Depends(get_current_user)):
    result = await uow.execute(select(Clue).where(Clue.id == clue_id))
    clue = result.scalar_one_or_none()
    if not clue:
        raise HTTPException(status_code=404, detail="Clue not found")
    for key, value in clue_data.model_dump(exclude_unset=True).items():
        setattr(clue, key, value)
    await uow.commit()
    await uow.refresh(clue)
    await write_audit_log(uow.session, current_user.id, "update", "clue", clue_id, {})
    return clue


@router.post("/{clue_id}/transfer")
async def transfer_clue(clue_id: UUID, body: ClueTransfer = Body(...), uow: UnitOfWork = Depends(get_uow), current_user: User = Depends(get_current_user)):
    result = await uow.execute(select(Clue).where(Clue.id == clue_id))
    clue = result.scalar_one_or_none()
    if not clue:
        raise HTTPException(status_code=404, detail="Clue not found")
    clue.status = "transferred"
    clue.transfer_target = body.target
    clue.transfer_date = func.now()
    clue.transfer_comment = body.comment
    await uow.commit()
    await write_audit_log(uow.session, current_user.id, "transfer", "clue", clue_id, {"target": body.target})
    return {"message": "Clue transferred"}
