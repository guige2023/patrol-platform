from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload
from typing import List, Optional
from uuid import UUID
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.dependencies import get_db, get_current_user
from app.models.rectification import Rectification
from app.models.user import User
from app.models.unit import Unit
from app.schemas.rectification import RectificationCreate, RectificationUpdate, RectificationResponse
from app.schemas.common import PaginatedResponse, PageResult
from app.core.audit import write_audit_log

router = APIRouter()

STATUS_LABELS = {
    "dispatched": "已派发",
    "signed": "已签收",
    "progressing": "整改中",
    "completed": "已完成",
    "submitted": "待验收",
    "verified": "已验收",
    "rejected": "已驳回",
}

ALERT_LABELS = {
    "green": "绿色",
    "yellow": "黄色",
    "red": "红色",
}


# ============================================================
# Special routes (MUST be before /{rect_id} to avoid route conflict)
# ============================================================

@router.get("/download")
async def export_rectifications(
    status: Optional[str] = None,
    alert_level: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export all rectifications as .xlsx."""
    query = (
        select(Rectification)
        .options(selectinload(Rectification.unit))
        .where(Rectification.is_active == True)
    )
    if status:
        query = query.where(Rectification.status == status)
    if alert_level:
        query = query.where(Rectification.alert_level == alert_level)
    query = query.order_by(Rectification.created_at.desc()).limit(10000)
    result = await db.execute(query)
    rects = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "整改记录"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1677FF")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = [
        "标题", "所属单位", "问题描述", "整改要求",
        "截止日期", "状态", "进度(%)", "预警级别",
        "签收日期", "完成日期", "整改报告", "验收意见",
        "创建时间",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for r in rects:
        ws.append([
            r.title or "",
            r.unit.name if r.unit else "",
            r.problem_description or "",
            r.rectification_requirement or "",
            r.deadline.strftime('%Y-%m-%d') if r.deadline else "",
            STATUS_LABELS.get(r.status, r.status or ""),
            r.progress or 0,
            ALERT_LABELS.get(r.alert_level, r.alert_level or ""),
            r.sign_date.strftime('%Y-%m-%d') if r.sign_date else "",
            r.completion_date.strftime('%Y-%m-%d') if r.completion_date else "",
            r.completion_report or "",
            r.verification_comment or "",
            r.created_at.strftime('%Y-%m-%d %H:%M') if r.created_at else "",
        ])

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''rectifications.xlsx"},
    )


@router.get("/template")
async def download_rectification_template(current_user: User = Depends(get_current_user)):
    """Download the rectification import template as .xlsx."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "整改记录导入模板"

    headers = [
        "标题", "单位ID", "问题描述", "整改要求",
        "截止日期(YYYY-MM-DD)", "预警级别(green/yellow/red)",
        "关联线索ID", "关联底稿ID",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    ws.append([
        "示例整改标题", "", "示例问题描述", "示例整改要求",
        "2026-06-30", "green", "", "",
    ])
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 25

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''rectification_template.xlsx"},
    )


@router.get("/export")
async def export_rectifications_by_year(
    year: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export rectifications as .xlsx (optionally filtered by year)."""
    query = select(Rectification).options(selectinload(Rectification.unit)).where(Rectification.is_active == True)
    if year:
        query = query.where(func.extract("year", Rectification.created_at) == year)
    query = query.order_by(Rectification.created_at.desc()).limit(10000)
    result = await db.execute(query)
    rects = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "整改记录导出"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1677FF")
    header_align = Alignment(horizontal="center", vertical="center")

    headers = ["标题", "单位", "问题描述", "整改要求", "截止日期", "状态", "进度(%)", "预警级别", "创建时间"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for r in rects:
        ws.append([
            r.title or "",
            r.unit.name if r.unit else "",
            r.problem_description or "",
            r.rectification_requirement or "",
            r.deadline.strftime('%Y-%m-%d') if r.deadline else "",
            STATUS_LABELS.get(r.status, r.status or ""),
            r.progress or 0,
            ALERT_LABELS.get(r.alert_level, r.alert_level or ""),
            r.created_at.strftime('%Y-%m-%d') if r.created_at else "",
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"rectifications_{year}.xlsx" if year else "rectifications_all.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


# ============================================================
# List / Create (MUST be before /{rect_id})
# ============================================================

@router.get("/", response_model=PaginatedResponse[RectificationResponse])
async def list_rectifications(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=9999),
    title: Optional[str] = None,
    status: Optional[str] = None,
    unit_id: Optional[UUID] = None,
    alert_level: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = select(Rectification).where(Rectification.is_active == True)
    if title:
        query = query.where(Rectification.title.ilike(f"%{title}%"))
    if status:
        query = query.where(Rectification.status == status)
    if unit_id:
        query = query.where(Rectification.unit_id == unit_id)
    if alert_level:
        query = query.where(Rectification.alert_level == alert_level)

    count_result = await db.execute(select(func.count()).select_from(query.subquery()))
    total = count_result.scalar()

    query = query.order_by(Rectification.created_at.desc()).offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(query)
    items = result.scalars().all()

    return PaginatedResponse(
        data=PageResult(items=items, total=total, page=page, page_size=page_size)
    )


@router.post("/import")
async def import_rectifications(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Import rectifications from .xlsx file."""
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            rows.append(dict(zip(headers, row)))

        imported = 0
        errors = []
        for i, row in enumerate(rows, start=2):
            try:
                deadline_str = row.get("截止日期(YYYY-MM-DD)")
                deadline = None
                if deadline_str:
                    try:
                        deadline = datetime.strptime(str(deadline_str), "%Y-%m-%d")
                    except Exception:
                        pass
                rect = Rectification(
                    title=row.get("标题", "导入整改") or "导入整改",
                    problem_description=row.get("问题描述", "导入问题") or "导入问题",
                    rectification_requirement=row.get("整改要求"),
                    deadline=deadline,
                    alert_level=row.get("预警级别(green/yellow/red)", "green"),
                    created_by=current_user.id,
                )
                db.add(rect)
                await db.flush()  # Flush to check for FK errors
                imported += 1
            except Exception as e:
                errors.append(f"Row {i}: {str(e)}")

        await db.commit()
        return {"message": f"Imported {imported} rectifications", "errors": errors[:20]}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Import failed: {str(e)}")


# ============================================================
# Batch operations (MUST be before /{rect_id})
# ============================================================

@router.post("/batch-delete")
async def batch_delete_rectifications(
    ids: List[UUID],
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Soft-delete multiple rectifications at once."""
    if not ids:
        raise HTTPException(status_code=400, detail="No IDs provided")
    result = await db.execute(
        select(Rectification).where(Rectification.id.in_(ids), Rectification.is_active == True)
    )
    rects = result.scalars().all()
    if not rects:
        raise HTTPException(status_code=404, detail="No rectifications found")
    for r in rects:
        r.is_active = False
    await db.commit()
    for r in rects:
        await write_audit_log(db, current_user.id, "delete", "rectification", r.id, {})
    return {"message": f"{len(rects)} rectifications deleted"}


@router.post("/batch-status")
async def batch_update_rectification_status(
    request: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Update status for multiple rectifications at once."""
    ids = request.get("ids", [])
    status = request.get("status")
    if not ids:
        raise HTTPException(status_code=400, detail="No IDs provided")
    if not status:
        raise HTTPException(status_code=400, detail="No status provided")
    valid_statuses = {"dispatched", "signed", "progressing", "completed", "submitted", "verified", "rejected"}
    if status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {', '.join(valid_statuses)}")
    result = await db.execute(
        select(Rectification).where(Rectification.id.in_(ids), Rectification.is_active == True)
    )
    rects = result.scalars().all()
    if not rects:
        raise HTTPException(status_code=404, detail="No rectifications found")
    old_statuses = {r.id: r.status for r in rects}
    for r in rects:
        r.status = status
    await db.commit()
    for r in rects:
        await write_audit_log(db, current_user.id, "update", "rectification", r.id,
                              {"old_status": old_statuses[r.id], "new_status": status})
    return {"message": f"{len(rects)} rectifications updated to '{status}'"}


# ============================================================
# Per-rectification routes (/{rect_id} and sub-routes)
# ============================================================

@router.post("/", response_model=RectificationResponse, status_code=201)
async def create_rectification(rect_data: RectificationCreate, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    rect = Rectification(**rect_data.model_dump(), created_by=current_user.id)
    db.add(rect)
    await db.commit()
    await db.refresh(rect)
    await write_audit_log(db, current_user.id, "create", "rectification", rect.id, {"title": rect.title})
    return rect


@router.get("/{rect_id}")
async def get_rectification(rect_id: UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(Rectification).where(Rectification.id == rect_id))
    rect = result.scalar_one_or_none()
    if not rect:
        raise HTTPException(status_code=404, detail="Rectification not found")
    return {"data": RectificationResponse.model_validate(rect), "message": "success"}


@router.put("/{rect_id}", response_model=RectificationResponse)
async def update_rectification(rect_id: UUID, rect_data: RectificationUpdate, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(Rectification).where(Rectification.id == rect_id))
    rect = result.scalar_one_or_none()
    if not rect:
        raise HTTPException(status_code=404, detail="Rectification not found")
    for key, value in rect_data.model_dump(exclude_unset=True).items():
        setattr(rect, key, value)
    await db.commit()
    await db.refresh(rect)
    await write_audit_log(db, current_user.id, "update", "rectification", rect_id, {})
    return rect


@router.patch("/{rect_id}/progress")
async def update_progress(rect_id: UUID, progress: int, details: Optional[List[dict]] = None, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(Rectification).where(Rectification.id == rect_id))
    rect = result.scalar_one_or_none()
    if not rect:
        raise HTTPException(status_code=404, detail="Rectification not found")
    rect.progress = max(0, min(100, progress))
    if details:
        rect.progress_details = details
    if rect.progress >= 100:
        rect.status = "completed"
        rect.completion_date = datetime.utcnow()
    await db.commit()
    await write_audit_log(db, current_user.id, "update_progress", "rectification", rect_id, {"progress": progress})
    return {"message": "Progress updated"}


@router.post("/{rect_id}/sign")
async def sign_rectification(rect_id: UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(Rectification).where(Rectification.id == rect_id))
    rect = result.scalar_one_or_none()
    if not rect:
        raise HTTPException(status_code=404, detail="Rectification not found")
    rect.status = "progressing"
    rect.sign_date = datetime.utcnow()
    rect.sign_by = current_user.id
    await db.commit()
    await write_audit_log(db, current_user.id, "sign", "rectification", rect_id, {})
    return {"message": "Rectification signed"}


@router.post("/{rect_id}/submit")
async def submit_rectification(rect_id: UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Submit a completed rectification for admin approval."""
    result = await db.execute(select(Rectification).where(Rectification.id == rect_id))
    rect = result.scalar_one_or_none()
    if not rect:
        raise HTTPException(status_code=404, detail="Rectification not found")
    if rect.status != "completed":
        raise HTTPException(status_code=400, detail="Only completed rectifications can be submitted")
    rect.status = "submitted"
    await db.commit()
    await write_audit_log(db, current_user.id, "submit", "rectification", rect_id, {})
    return {"message": "Rectification submitted for approval"}


@router.post("/{rect_id}/verify")
async def verify_rectification(rect_id: UUID, comment: Optional[str] = None, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(Rectification).where(Rectification.id == rect_id))
    rect = result.scalar_one_or_none()
    if not rect:
        raise HTTPException(status_code=404, detail="Rectification not found")
    rect.status = "verified"
    rect.verified_by = current_user.id
    rect.verified_at = datetime.utcnow()
    rect.verification_comment = comment
    await db.commit()
    await write_audit_log(db, current_user.id, "verify", "rectification", rect_id, {})
    return {"message": "Rectification verified"}


@router.post("/{rect_id}/confirm")
async def confirm_rectification(
    rect_id: UUID,
    body: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Confirm a rectification as completed or rejected."""
    result = await db.execute(select(Rectification).where(Rectification.id == rect_id))
    rect = result.scalar_one_or_none()
    if not rect:
        raise HTTPException(status_code=404, detail="Rectification not found")
    is_completed = body.get("is_completed")
    notes = body.get("notes")
    rect.confirmed_completed = is_completed
    rect.confirm_notes = notes
    rect.confirmed_at = datetime.utcnow()
    rect.confirmed_by = current_user.id
    await db.commit()
    await write_audit_log(db, current_user.id, "confirm", "rectification", rect_id,
                          {"is_completed": is_completed, "notes": notes})
    return {"message": "Rectification confirmed"}


@router.get("/{rect_id}/export-pdf")
async def export_rectification_pdf(
    rect_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export a single rectification as .xlsx (PDF requires reportlab)."""
    result = await db.execute(
        select(Rectification)
        .options(selectinload(Rectification.unit))
        .where(Rectification.id == rect_id)
    )
    rect = result.scalar_one_or_none()
    if not rect:
        raise HTTPException(status_code=404, detail="Rectification not found")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "整改通知书"

    ws.append(["巡察整改通知书"])
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:D1")

    ws.append([])
    fields = [
        ("标题", rect.title or ""),
        ("被整改单位", rect.unit.name if rect.unit else ""),
        ("问题描述", rect.problem_description or ""),
        ("整改要求", rect.rectification_requirement or ""),
        ("截止日期", rect.deadline.strftime("%Y-%m-%d") if rect.deadline else ""),
        ("当前状态", STATUS_LABELS.get(rect.status, rect.status or "")),
        ("完成情况", "已完成" if rect.confirmed_completed else "未完成"),
        ("确认意见", rect.confirm_notes or ""),
    ]
    for label, value in fields:
        ws.append([label, value])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''rectification_{rect_id}.xlsx"},
    )


@router.post("/{rect_id}/reimport")
async def reimport_rectification(
    rect_id: UUID,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Reimport/update a rectification from .xlsx file."""
    result = await db.execute(select(Rectification).where(Rectification.id == rect_id))
    rect = result.scalar_one_or_none()
    if not rect:
        raise HTTPException(status_code=404, detail="Rectification not found")

    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            data = dict(zip(headers, row))
            if data.get("标题"):
                rect.title = data["标题"]
            if data.get("问题描述"):
                rect.problem_description = data["问题描述"]
            if data.get("整改要求"):
                rect.rectification_requirement = data["整改要求"]
            break

        await db.commit()
        await write_audit_log(db, current_user.id, "reimport", "rectification", rect_id, {})
        return {"message": "Rectification reimported successfully"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Reimport failed: {str(e)}")


@router.delete("/{rect_id}")
async def delete_rectification(rect_id: UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Soft-delete a rectification by setting is_active=False."""
    result = await db.execute(select(Rectification).where(Rectification.id == rect_id))
    rect = result.scalar_one_or_none()
    if not rect:
        raise HTTPException(status_code=404, detail="Rectification not found")
    rect.is_active = False
    await db.commit()
    await write_audit_log(db, current_user.id, "delete", "rectification", rect_id, {})
    return {"message": "Rectification deleted"}
