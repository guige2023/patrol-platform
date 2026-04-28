from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload
from typing import List, Optional
from uuid import UUID
import io, codecs

from app.dependencies import get_uow, get_current_user
from app.database import UnitOfWork
from app.models.unit import Unit
from app.models.user import User
from app.schemas.unit import UnitCreate, UnitUpdate, UnitResponse, UnitTreeResponse
from app.schemas.common import PaginatedResponse, PageResult
from app.core.audit import write_audit_log
import json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

router = APIRouter()


def build_unit_tree(units: List[Unit], parent_id: Optional[UUID] = None) -> List[UnitTreeResponse]:
    tree = []
    for unit in units:
        if unit.parent_id == parent_id:
            children = build_unit_tree(units, unit.id)
            resp = UnitTreeResponse(
                id=unit.id,
                name=unit.name,
                org_code=unit.org_code,
                parent_id=unit.parent_id,
                unit_type=unit.unit_type,
                level=unit.level,
                sort_order=unit.sort_order,
                tags=unit.tags or {},
                business_tags=unit.business_tags or [],
                profile=unit.profile,
                leadership=unit.leadership,
                contact=unit.contact,
                is_active=unit.is_active,
                created_at=unit.created_at,
                children=children,
            )
            tree.append(resp)
    return tree


@router.get("/tree", response_model=List[UnitTreeResponse])
async def get_unit_tree(
    uow: UnitOfWork = Depends(get_uow),
    current_user: User = Depends(get_current_user),
):
    result = await uow.execute(select(Unit).where(Unit.is_active == True).order_by(Unit.sort_order))
    units = result.scalars().all()
    return build_unit_tree(list(units))


@router.get("/", response_model=PaginatedResponse[UnitResponse])
async def list_units(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=9999),
    name: Optional[str] = None,
    unit_type: Optional[str] = None,
    parent_id: Optional[UUID] = None,
    uow: UnitOfWork = Depends(get_uow),
    current_user: User = Depends(get_current_user),
):
    query = select(Unit)
    if name:
        query = query.where(Unit.name.ilike(f"%{name}%"))
    if unit_type:
        query = query.where(Unit.unit_type == unit_type)
    if parent_id:
        query = query.where(Unit.parent_id == parent_id)
    
    count_result = await uow.execute(select(func.count()).select_from(query.subquery()))
    total = count_result.scalar()
    
    query = query.order_by(Unit.sort_order).offset((page - 1) * page_size).limit(page_size)
    result = await uow.execute(query)
    items = result.scalars().all()
    
    return PaginatedResponse(
        data=PageResult(items=items, total=total, page=page, page_size=page_size)
    )


@router.get("/export")
async def export_units(
    name: Optional[str] = None,
    unit_type: Optional[str] = None,
    uow: UnitOfWork = Depends(get_uow),
    current_user: User = Depends(get_current_user),
):
    """Export units as real .xlsx file."""
    query = select(Unit).where(Unit.is_active == True)
    if name:
        query = query.where(Unit.name.ilike(f"%{name}%"))
    if unit_type:
        query = query.where(Unit.unit_type == unit_type)
    query = query.order_by(Unit.sort_order or 0, Unit.created_at.desc()).limit(10000)
    result = await uow.execute(query)
    units = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "单位档案"

    # Style helpers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1677FF")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = ["单位名称", "组织编码", "单位类型", "单位级别", "排序", "标签(逗号分隔)", "简介", "最近巡察年份", "巡察历史", "是否可用"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Sample data to show expected format (only for empty export)
    sample = [
        ["示例单位A", "ORG001", "党委", "一级单位", 1, "优秀,党建", "示例简介", 2024, "2021年第一轮", "是"],
        ["示例单位B", "ORG002", "政府", "二级单位", 2, "业务骨干", "示例简介2", 2023, "", "是"],
    ]
    if not units:
        for row in sample:
            ws.append(row)

    for u in units:
        tags_val = u.tags
        if isinstance(tags_val, dict):
            tags_str = ",".join(str(v) for v in tags_val.values() if v)
        elif isinstance(tags_val, list):
            tags_str = ",".join(str(v) for v in tags_val)
        else:
            tags_str = str(tags_val) if tags_val else ""
        ws.append([
            u.name or "",
            u.org_code or "",
            u.unit_type or "",
            u.level or "",
            u.sort_order or "",
            tags_str,
            u.profile or "",
            u.last_inspection_year or "",
            u.inspection_history or "",
            "是" if u.is_active else "否",
        ])

    # Auto column width
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''units_export.xlsx"},
    )


@router.get("/template")
async def download_unit_template(
    current_user: User = Depends(get_current_user),
):
    """Download unit import template (.xlsx)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "单位档案导入模板"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="52C41A")
    header_align = Alignment(horizontal="center", vertical="center")
    note_fill = PatternFill("solid", fgColor="FFF7E6")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = ["单位名称*", "组织编码*", "单位类型", "单位级别", "排序", "标签(逗号分隔)", "简介", "最近巡察年份", "巡察历史", "是否可用"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Notes row
    notes = [
        "必填", "唯一标识", "党委/纪委/组织部/宣传部/政府/其他", "一级单位/二级单位", "数字，越小越靠前", "多个用逗号分隔", "单位简介", "如 2024", "如 2021年第一轮", "是/否，默认为是",
    ]
    ws.append(notes)
    for cell in ws[2]:
        cell.fill = note_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        cell.font = Font(size=10, color="999999")

    # Sample data
    sample = [
        ["示例单位A", "ORG001", "党委", "一级单位", 1, "优秀,党建", "示例简介", 2024, "2021年第一轮", "是"],
        ["示例单位B", "ORG002", "政府", "二级单位", 2, "业务骨干", "示例简介2", 2023, "", "是"],
    ]
    for row in sample:
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[2].height = 36
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 6, 45)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''unit_template.xlsx"},
    )


@router.get("/{unit_id}")
async def get_unit(unit_id: UUID, uow: UnitOfWork = Depends(get_uow), current_user: User = Depends(get_current_user)):
    result = await uow.execute(select(Unit).where(Unit.id == unit_id))
    unit = result.scalar_one_or_none()
    if not unit:
        raise HTTPException(status_code=404, detail="Unit not found")
    return {"data": UnitResponse.model_validate(unit), "message": "success"}


@router.post("/", response_model=UnitResponse, status_code=201)
async def create_unit(unit_data: UnitCreate, uow: UnitOfWork = Depends(get_uow), current_user: User = Depends(get_current_user)):
    unit = Unit(**unit_data.model_dump())
    uow.add(unit)
    await uow.commit()
    await uow.refresh(unit)
    await write_audit_log(uow.session, current_user.id, "create", "unit", unit.id, {"name": unit.name})
    return unit


@router.put("/{unit_id}", response_model=UnitResponse)
async def update_unit(unit_id: UUID, unit_data: UnitUpdate, uow: UnitOfWork = Depends(get_uow), current_user: User = Depends(get_current_user)):
    result = await uow.execute(select(Unit).where(Unit.id == unit_id))
    unit = result.scalar_one_or_none()
    if not unit:
        raise HTTPException(status_code=404, detail="Unit not found")
    for key, value in unit_data.model_dump(exclude_unset=True).items():
        setattr(unit, key, value)
    await uow.commit()
    await uow.refresh(unit)
    await write_audit_log(uow.session, current_user.id, "update", "unit", unit.id, {"name": unit.name})
    return unit


@router.delete("/{unit_id}")
async def delete_unit(unit_id: UUID, uow: UnitOfWork = Depends(get_uow), current_user: User = Depends(get_current_user)):
    result = await uow.execute(select(Unit).where(Unit.id == unit_id))
    unit = result.scalar_one_or_none()
    if not unit:
        raise HTTPException(status_code=404, detail="Unit not found")
    unit.is_active = False
    await uow.commit()
    await write_audit_log(uow.session, current_user.id, "delete", "unit", unit.id, {"name": unit.name})
    return {"message": "Unit deleted"}



@router.post("/import")
async def import_units(
    file: UploadFile = File(...),
    uow: UnitOfWork = Depends(get_uow),
    current_user: User = Depends(get_current_user),
):
    """
    导入单位数据（Excel .xlsx）
    必填列：name, org_code
    可选列：parent_id, unit_type, level, sort_order, tags, profile, leadership( JSON),
            contact( JSON), is_active
    """
    import json, openpyxl
    from app.schemas.unit import UnitCreate

    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 文件")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    # 优先用用户当前激活的 sheet，若其首行无内容则回退到第一个有数据的 sheet
    ws = wb.active
    headers = [str(h.value).strip() if h.value else "" for h in ws[1]]
    if not any(headers):
        for sheet in wb.worksheets:
            candidate = [str(h.value).strip() if h.value else "" for h in sheet[1]]
            if any(candidate):
                ws = sheet
                headers = candidate
                break

    # 中文表头映射
    HEADER_ALIASES = {
        "单位名称*": "name", "单位名称": "name",
        "组织编码*": "org_code", "组织编码": "org_code",
        "单位类型": "unit_type", "类型": "unit_type",
        "单位级别": "level", "级别": "level",
        "排序": "sort_order", "顺序": "sort_order",
        "标签(逗号分隔)": "tags", "标签": "tags",
        "简介": "profile",
        "最近巡察年份": "last_inspection_year", "巡察年份": "last_inspection_year",
        "巡察历史": "inspection_history",
        "是否可用": "is_active",
        "上级单位": "parent_id",
        "领导信息(JSON)": "leadership",
        "联系方式(JSON)": "contact",
    }

    # 标准化列名（去除*后缀和空格）
    # 先建立一个 原始表头 → 映射名 的反向索引，支持部分匹配
    header_to_key = {}
    for orig_key, mapped_key in HEADER_ALIASES.items():
        header_to_key[orig_key] = mapped_key

    col_map = {}
    for i, h in enumerate(headers):
        h_stripped = h.strip().rstrip("*").strip()
        # 精确匹配
        key = header_to_key.get(h_stripped)
        if key is None:
            # 尝试用原始 HEADER_ALIASES key 的前缀匹配（如 "单位类型/类型" → "单位类型"）
            for alias_src, mapped_key in HEADER_ALIASES.items():
                if h_stripped.startswith(alias_src.strip().rstrip("*").strip()):
                    key = mapped_key
                    break
            else:
                key = h_stripped.lower()
        col_map[key] = i

    required = ["name", "org_code"]
    for col in required:
        if col not in col_map:
            raise HTTPException(status_code=400, detail=f"缺少必填列: {col}")

    # 加载字段选项配置，构建校验规则
    from app.models.field_option import FieldOption
    field_opts_result = await uow.execute(select(FieldOption))
    all_field_opts = field_opts_result.scalars().all()
    # field_key → set of valid labels
    field_option_map: dict[str, set[str]] = {}
    for fo in all_field_opts:
        field_option_map[fo.field_key] = {opt["value"] for opt in json.loads(fo.options) if isinstance(opt, dict) and opt.get("value")}

    # 需要校验的字段映射（表头 → field_key）
    validated_fields = {
        "unit_type": "unit_type",
        "level": "unit_level",
    }

    created, skipped, errors = 0, 0, []
    first_id = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            name = row[col_map["name"]]
            org_code = row[col_map["org_code"]]
            if not name or not org_code:
                errors.append(f"第{row_idx}行: name 或 org_code 为空")
                continue

            # 检查是否已存在
            exist = await uow.execute(
                select(Unit).where(Unit.org_code == str(org_code))
            )
            if exist.scalar_one_or_none():
                skipped += 1
                continue

            data = {
                "name": str(name).strip(),
                "org_code": str(org_code).strip(),
                "unit_type": _cell(row, col_map, "unit_type"),
                "level": _cell(row, col_map, "level"),
                "sort_order": _int_cell(row, col_map, "sort_order") or 0,
                "tags": _list_cell(row, col_map, "tags"),
                "profile": _cell(row, col_map, "profile"),
                "leadership": _json_cell(row, col_map, "leadership"),
                "contact": _json_cell(row, col_map, "contact"),
                "is_active": _bool_cell(row, col_map, "is_active", default=True),
                "parent_id": None,
                "last_inspection_year": _int_cell(row, col_map, "last_inspection_year"),
                "inspection_history": _cell(row, col_map, "inspection_history"),
            }

            # 字段选项范围校验
            for col_key, field_key in validated_fields.items():
                # 只对Excel里有这列 且 field_key 有配置的情况下校验
                if col_key in col_map and field_key in field_option_map:
                    raw_val = data.get(col_key)
                    if raw_val:  # 只校验有值的
                        valid_labels = field_option_map[field_key]
                        if raw_val not in valid_labels:
                            errors.append(
                                f"第{row_idx}行 [{name}]: {col_key}='{raw_val}' 不在系统允许的范围内。"
                                f"可选值: {', '.join(sorted(valid_labels))}"
                            )

            # 如果有校验错误，跳过此行但不中断整批
            # 找到第一个校验错误对应的行
            if any(f"第{row_idx}行" in e for e in errors):
                continue

            unit = Unit(**data)
            uow.add(unit)
            await uow.flush()
            if first_id is None:
                first_id = unit.id
            created += 1
        except Exception as e:
            errors.append(f"第{row_idx}行: {str(e)}")

    await uow.commit()
    # 只要有校验错误就返回详细提示
    validation_errors = [e for e in errors if "不在系统允许的范围内" in e or "不在可选值范围内" in e]
    if validation_errors:
        raise HTTPException(
            status_code=400,
            detail={
                "message": f"有 {len(validation_errors)} 条数据校验不通过，请修改后重新导入",
                "errors": validation_errors,
                "created": created,
                "skipped": skipped,
            }
        )

    if errors and created == 0:
        raise HTTPException(status_code=400, detail=f"导入失败: {'; '.join(errors[:5])}")

    if first_id is not None:
        await write_audit_log(uow.session, current_user.id, "import", "unit", first_id, {
            "created": created, "skipped": skipped
        })
    return {
        "message": f"导入完成：新增 {created} 条，{skipped} 条因组织编码重复被跳过",
        "detail": f"{skipped}条记录因组织编码重复被跳过，可前往列表手动编辑覆盖",
        "created": created,
        "skipped": skipped,
        "errors": errors[:10],
    } if created > 0 else {
        "message": f"无新数据导入（{skipped}条组织编码重复）",
        "detail": f"{skipped}条记录因组织编码重复被跳过，可前往列表手动编辑覆盖",
        "created": 0,
        "skipped": skipped,
        "errors": errors[:10],
    }


def _cell(row, col_map, key):
    if key not in col_map:
        return None
    v = row[col_map[key]]
    return str(v).strip() if v is not None else None


def _int_cell(row, col_map, key):
    v = _cell(row, col_map, key)
    try:
        return int(v) if v else None
    except (ValueError, TypeError):
        return None


def _bool_cell(row, col_map, key, default=False):
    v = _cell(row, col_map, key)
    if v is None:
        return default
    return str(v).lower() in ("true", "1", "是", "yes")


def _list_cell(row, col_map, key):
    v = _cell(row, col_map, key)
    if not v:
        return []
    return [s.strip() for s in str(v).split(",") if s.strip()]


def _json_cell(row, col_map, key):
    import json
    v = _cell(row, col_map, key)
    if not v:
        return None
    try:
        return json.loads(v)
    except (json.JSONDecodeError, TypeError):
        return {"raw": v}
